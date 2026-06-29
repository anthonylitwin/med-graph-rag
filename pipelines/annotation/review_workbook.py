from __future__ import annotations

import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_REVIEW_SHEETS = [
    "documents",
    "chunks",
    "gold_entities",
    "gold_relationships",
    "rejected_candidates",
    "annotation_notes",
]


@dataclass(frozen=True)
class ReviewWorkbook:
    path: Path
    documents: list[dict[str, Any]]
    chunks: list[dict[str, Any]]
    gold_entities: list[dict[str, Any]]
    gold_relationships: list[dict[str, Any]]
    rejected_candidates: list[dict[str, Any]]
    annotation_notes: list[dict[str, Any]]


def _sheet_rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(header or "") for header in rows[0]]
    records: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(value is not None for value in row):
            continue
        record = {header: value for header, value in zip(headers, row) if header}
        record["_row_number"] = row_number
        records.append(record)
    return records


def read_review_workbook(workbook_path: Path) -> ReviewWorkbook:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Annotation workbook does not exist: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        missing = [sheet for sheet in REQUIRED_REVIEW_SHEETS if sheet not in workbook.sheetnames]
        if missing:
            raise ValueError(f"Annotation workbook is missing required sheet(s): {', '.join(missing)}")

        return ReviewWorkbook(
            path=workbook_path,
            documents=_sheet_rows(workbook, "documents"),
            chunks=_sheet_rows(workbook, "chunks"),
            gold_entities=_sheet_rows(workbook, "gold_entities"),
            gold_relationships=_sheet_rows(workbook, "gold_relationships"),
            rejected_candidates=_sheet_rows(workbook, "rejected_candidates"),
            annotation_notes=_sheet_rows(workbook, "annotation_notes"),
        )
    finally:
        workbook.close()


def copy_reviewed_workbook(source_path: Path, destination_path: Path) -> Path:
    destination_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, destination_path)
    return destination_path


__all__ = ["REQUIRED_REVIEW_SHEETS", "ReviewWorkbook", "copy_reviewed_workbook", "read_review_workbook"]
