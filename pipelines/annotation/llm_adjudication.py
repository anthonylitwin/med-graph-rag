from __future__ import annotations

import hashlib
import json
from dataclasses import asdict, dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from packages.llm.profiles import resolve_model_profile
from packages.llm.providers import LanguageModel, ModelCallRecord, get_language_model
from pipelines.annotation.workbook import ALLOWED_VALUES, ENTITY_HEADERS, RELATIONSHIP_HEADERS


PROMPT_VERSION = "annotation_adjudication_v001"
ENTITY_ACCEPT_DECISIONS = {"accepted", "accepted_with_normalization", "revised", "added"}
ENTITY_REJECT_DECISIONS = {"rejected", "superseded"}
RELATIONSHIP_ACCEPT_DECISIONS = {"accepted", "revised", "replaced", "added"}
RELATIONSHIP_REJECT_DECISIONS = {"rejected"}


@dataclass(frozen=True)
class ChunkReviewResult:
    chunk_id: str
    pmcid: str
    status: str
    entity_audit_path: str
    relationship_audit_path: str
    error: str = ""

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class LlmAdjudicationResult:
    llm_review: bool
    model_profile: dict[str, str]
    provider: str
    model: str
    reviewed_chunk_count: int
    failed_chunk_count: int
    model_call_paths: list[str]
    chunks: list[ChunkReviewResult]

    def to_dict(self) -> dict[str, Any]:
        return {
            "llm_review": self.llm_review,
            "model_profile": self.model_profile,
            "provider": self.provider,
            "model": self.model,
            "reviewed_chunk_count": self.reviewed_chunk_count,
            "failed_chunk_count": self.failed_chunk_count,
            "model_call_paths": self.model_call_paths,
            "chunks": [chunk.to_dict() for chunk in self.chunks],
        }


def _now_iso() -> str:
    return datetime.now(UTC).isoformat()


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _hash_id(prefix: str, *parts: Any) -> str:
    payload = "|".join(str(part or "") for part in parts)
    return f"{prefix}:{hashlib.sha1(payload.encode('utf-8')).hexdigest()[:16]}"


def _headers(ws: Any) -> dict[str, int]:
    return {str(cell.value or ""): cell.column for cell in ws[1] if cell.value}


def _row_dict(ws: Any, header_map: dict[str, int], row_number: int) -> dict[str, Any]:
    row = {header: ws.cell(row=row_number, column=column).value for header, column in header_map.items()}
    row["_row_number"] = row_number
    return row


def _rows(ws: Any, header_map: dict[str, int]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row_number in range(2, ws.max_row + 1):
        row = _row_dict(ws, header_map, row_number)
        if any(value is not None for key, value in row.items() if key != "_row_number"):
            records.append(row)
    return records


def _cell_set(ws: Any, header_map: dict[str, int], row_number: int, header: str, value: Any) -> None:
    if header in header_map:
        ws.cell(row=row_number, column=header_map[header]).value = value


def _set_if_present(ws: Any, header_map: dict[str, int], row_number: int, header: str, value: Any) -> None:
    if _clean(value):
        _cell_set(ws, header_map, row_number, header, value)


def _append_row(ws: Any, headers: list[str], values: dict[str, Any]) -> None:
    ws.append([values.get(header, "") for header in headers])


def _field(row: dict[str, Any], field: str) -> str:
    return _clean(row.get(field))


def _absolute_offsets_match(text: str, evidence: str, start_value: Any, end_value: Any, chunk_start_value: Any) -> bool:
    try:
        start = int(start_value)
        end = int(end_value)
        chunk_start = int(chunk_start_value or 0)
    except (TypeError, ValueError):
        return False
    local_start = start - chunk_start
    local_end = end - chunk_start
    if local_start < 0 or local_end < local_start or local_end > len(text):
        return False
    return text[local_start:local_end].casefold() == evidence.casefold()


def _evidence_supported(chunk: dict[str, Any], evidence: Any, start: Any = "", end: Any = "") -> bool:
    evidence_text = _clean(evidence)
    chunk_text = _field(chunk, "chunk_text")
    if not evidence_text or not chunk_text:
        return False
    return evidence_text.casefold() in chunk_text.casefold() or _absolute_offsets_match(
        chunk_text,
        evidence_text,
        start,
        end,
        chunk.get("start_char"),
    )


def _model_call_path(review_root: Path, chunk: dict[str, Any], stage: str) -> Path:
    pmcid = _field(chunk, "pmcid") or "unknown_pmcid"
    chunk_id = _field(chunk, "chunk_id") or "unknown_chunk"
    return review_root / "model_calls" / pmcid / f"{chunk_id}.{stage}.json"


def _write_model_call(path: Path, record: ModelCallRecord) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(record.to_dict(), indent=2), encoding="utf-8")
    return path.as_posix()


def _frontier_language_model(model_profile: str, model: str | None) -> tuple[dict[str, str], LanguageModel]:
    profile = resolve_model_profile(model_profile, extractor_model=model)
    if profile.name != "frontier" or profile.extractor_provider != "openai":
        raise ValueError("LLM adjudication only supports the frontier OpenAI model profile.")
    language_model = get_language_model("openai", profile.extractor_model)
    if getattr(language_model, "provider", "") != "openai":
        raise ValueError("LLM adjudication only supports the OpenAI frontier provider.")
    return profile.to_dict(), language_model


def _entity_review_schema() -> dict[str, Any]:
    string = {"type": "string"}
    return {
        "type": "json_schema",
        "name": "medgraphrag_entity_adjudication",
        "strict": True,
        "schema": {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "entity_reviews": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "additionalProperties": False,
                        "properties": {
                            "entity_gold_id": string,
                            "decision": {
                                "type": "string",
                                "enum": sorted(ENTITY_ACCEPT_DECISIONS | ENTITY_REJECT_DECISIONS),
                            },
                            "revised_entity_type": {"type": "string", "enum": ["", *ALLOWED_VALUES["entity_type"]]},
                            "revised_entity_text": string,
                            "revised_normalized_name": string,
                            "evidence_text": string,
                            "evidence_start_char": string,
                            "evidence_end_char": string,
                            "reason": string,
                        },
                        "required": [
                            "entity_gold_id",
                            "decision",
                            "revised_entity_type",
                            "revised_entity_text",
                            "revised_normalized_name",
                            "evidence_text",
                            "evidence_start_char",
                            "evidence_end_char",
                            "reason",
                        ],
                    },
                },
                "entity_additions": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "additionalProperties": False,
                        "properties": {
                            "entity_gold_id": string,
                            "entity_type": {"type": "string", "enum": ALLOWED_VALUES["entity_type"]},
                            "entity_text": string,
                            "normalized_name": string,
                            "evidence_text": string,
                            "evidence_start_char": string,
                            "evidence_end_char": string,
                            "reason": string,
                        },
                        "required": [
                            "entity_gold_id",
                            "entity_type",
                            "entity_text",
                            "normalized_name",
                            "evidence_text",
                            "evidence_start_char",
                            "evidence_end_char",
                            "reason",
                        ],
                    },
                },
                "review_notes": string,
            },
            "required": ["entity_reviews", "entity_additions", "review_notes"],
        },
    }


def _relationship_review_schema() -> dict[str, Any]:
    string = {"type": "string"}
    yes_no_unclear = ["yes", "no", "unclear"]
    return {
        "type": "json_schema",
        "name": "medgraphrag_relationship_adjudication",
        "strict": True,
        "schema": {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "relationship_reviews": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "additionalProperties": False,
                        "properties": {
                            "relationship_gold_id": string,
                            "decision": {
                                "type": "string",
                                "enum": sorted(RELATIONSHIP_ACCEPT_DECISIONS | RELATIONSHIP_REJECT_DECISIONS),
                            },
                            "relationship_type": {"type": "string", "enum": ["", *ALLOWED_VALUES["relationship_type"]]},
                            "source_entity_id": string,
                            "source_entity_type": {"type": "string", "enum": ["", *ALLOWED_VALUES["entity_type"]]},
                            "source_entity_text": string,
                            "source_normalized_name": string,
                            "target_entity_id": string,
                            "target_entity_type": {"type": "string", "enum": ["", *ALLOWED_VALUES["entity_type"]]},
                            "target_entity_text": string,
                            "target_normalized_name": string,
                            "evidence_text": string,
                            "evidence_start_char": string,
                            "evidence_end_char": string,
                            "direction_verified": {"type": "string", "enum": ALLOWED_VALUES["direction_verified"]},
                            "explicit_or_implied": {"type": "string", "enum": ALLOWED_VALUES["explicit_or_implied"]},
                            "negated": {"type": "string", "enum": yes_no_unclear},
                            "speculative": {"type": "string", "enum": yes_no_unclear},
                            "reason": string,
                        },
                        "required": [
                            "relationship_gold_id",
                            "decision",
                            "relationship_type",
                            "source_entity_id",
                            "source_entity_type",
                            "source_entity_text",
                            "source_normalized_name",
                            "target_entity_id",
                            "target_entity_type",
                            "target_entity_text",
                            "target_normalized_name",
                            "evidence_text",
                            "evidence_start_char",
                            "evidence_end_char",
                            "direction_verified",
                            "explicit_or_implied",
                            "negated",
                            "speculative",
                            "reason",
                        ],
                    },
                },
                "relationship_additions": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "additionalProperties": False,
                        "properties": {
                            "relationship_gold_id": string,
                            "relationship_type": {"type": "string", "enum": ALLOWED_VALUES["relationship_type"]},
                            "source_entity_id": string,
                            "source_entity_type": {"type": "string", "enum": ALLOWED_VALUES["entity_type"]},
                            "source_entity_text": string,
                            "source_normalized_name": string,
                            "target_entity_id": string,
                            "target_entity_type": {"type": "string", "enum": ALLOWED_VALUES["entity_type"]},
                            "target_entity_text": string,
                            "target_normalized_name": string,
                            "evidence_text": string,
                            "evidence_start_char": string,
                            "evidence_end_char": string,
                            "direction_verified": {"type": "string", "enum": ALLOWED_VALUES["direction_verified"]},
                            "explicit_or_implied": {"type": "string", "enum": ALLOWED_VALUES["explicit_or_implied"]},
                            "negated": {"type": "string", "enum": yes_no_unclear},
                            "speculative": {"type": "string", "enum": yes_no_unclear},
                            "reason": string,
                        },
                        "required": [
                            "relationship_gold_id",
                            "relationship_type",
                            "source_entity_id",
                            "source_entity_type",
                            "source_entity_text",
                            "source_normalized_name",
                            "target_entity_id",
                            "target_entity_type",
                            "target_entity_text",
                            "target_normalized_name",
                            "evidence_text",
                            "evidence_start_char",
                            "evidence_end_char",
                            "direction_verified",
                            "explicit_or_implied",
                            "negated",
                            "speculative",
                            "reason",
                        ],
                    },
                },
                "review_notes": string,
            },
            "required": ["relationship_reviews", "relationship_additions", "review_notes"],
        },
    }


def _entity_prompt(chunk: dict[str, Any], entities: list[dict[str, Any]]) -> str:
    payload = {
        "task": "ENTITY_ADJUDICATION",
        "rules": [
            "Use only the chunk text as evidence.",
            "Return one entity_reviews item for every candidate entity row.",
            "Accept only standalone biomedical concepts supported by the chunk.",
            "Reject unsupported, generic, duplicate, wrong-type, or modifier-only entities.",
            "Use entity_additions only for important missing entities explicitly present in the chunk.",
        ],
        "ontology": {"entity_types": ALLOWED_VALUES["entity_type"]},
        "chunk": {
            "chunk_id": _field(chunk, "chunk_id"),
            "document_id": _field(chunk, "document_id"),
            "pmcid": _field(chunk, "pmcid"),
            "section": _field(chunk, "chunk_section"),
            "start_char": _field(chunk, "start_char"),
            "end_char": _field(chunk, "end_char"),
            "text": _field(chunk, "chunk_text"),
        },
        "candidate_entities": [
            {header: row.get(header, "") for header in ENTITY_HEADERS if header in row}
            for row in entities
        ],
    }
    return (
        "You are reviewing extracted biomedical entities for a MedGraphRAG gold dataset.\n"
        "Return strict JSON matching the supplied schema. Do not include prose outside JSON.\n\n"
        f"Input JSON:\n{json.dumps(payload, indent=2, ensure_ascii=True)}\n"
    )


def _relationship_prompt(
    chunk: dict[str, Any],
    entities: list[dict[str, Any]],
    relationships: list[dict[str, Any]],
) -> str:
    payload = {
        "task": "RELATIONSHIP_ADJUDICATION",
        "rules": [
            "Use only the chunk text as evidence.",
            "Review relationships only against accepted entity rows.",
            "Return one relationship_reviews item for every candidate relationship row.",
            "Reject relationships with rejected or unsupported endpoints.",
            "Accepted relationships must have direction_verified=yes, negated=no, and speculative=no.",
        ],
        "ontology": {
            "entity_types": ALLOWED_VALUES["entity_type"],
            "relationship_types": ALLOWED_VALUES["relationship_type"],
        },
        "chunk": {
            "chunk_id": _field(chunk, "chunk_id"),
            "document_id": _field(chunk, "document_id"),
            "pmcid": _field(chunk, "pmcid"),
            "section": _field(chunk, "chunk_section"),
            "start_char": _field(chunk, "start_char"),
            "end_char": _field(chunk, "end_char"),
            "text": _field(chunk, "chunk_text"),
        },
        "accepted_entities": [
            {header: row.get(header, "") for header in ENTITY_HEADERS if header in row}
            for row in entities
            if _field(row, "annotation_status") == "accepted"
        ],
        "candidate_relationships": [
            {header: row.get(header, "") for header in RELATIONSHIP_HEADERS if header in row}
            for row in relationships
        ],
    }
    return (
        "You are reviewing extracted biomedical relationships for a MedGraphRAG gold dataset.\n"
        "Return strict JSON matching the supplied schema. Do not include prose outside JSON.\n\n"
        f"Input JSON:\n{json.dumps(payload, indent=2, ensure_ascii=True)}\n"
    )


def _call_json_model(
    language_model: LanguageModel,
    prompt: str,
    schema: dict[str, Any],
    audit_path: Path,
) -> tuple[dict[str, Any], str, str]:
    record = language_model.generate_json_record(prompt, schema, prompt_version=PROMPT_VERSION)
    path = _write_model_call(audit_path, record)
    if record.status != "ok":
        return {}, path, record.error or "model call failed"
    return record.parsed_json, path, ""


def _validate_entity_response(
    response: dict[str, Any],
    chunk: dict[str, Any],
    existing_rows: list[dict[str, Any]],
) -> list[str]:
    errors: list[str] = []
    reviews = response.get("entity_reviews")
    additions = response.get("entity_additions")
    if not isinstance(reviews, list):
        return ["entity_reviews must be an array"]
    if not isinstance(additions, list):
        return ["entity_additions must be an array"]

    expected_ids = {_field(row, "entity_gold_id") for row in existing_rows}
    observed_ids = {_clean(item.get("entity_gold_id")) for item in reviews if isinstance(item, dict)}
    if observed_ids != expected_ids:
        missing = sorted(expected_ids - observed_ids)
        extra = sorted(observed_ids - expected_ids)
        errors.append(f"entity_reviews must cover exactly the chunk entity IDs; missing={missing}; extra={extra}")

    existing_by_id = {_field(row, "entity_gold_id"): row for row in existing_rows}
    for item in reviews:
        if not isinstance(item, dict):
            errors.append("entity review item is not an object")
            continue
        row = existing_by_id.get(_clean(item.get("entity_gold_id")))
        decision = _clean(item.get("decision"))
        if decision not in ENTITY_ACCEPT_DECISIONS | ENTITY_REJECT_DECISIONS:
            errors.append(f"invalid entity decision: {decision}")
        if decision in ENTITY_ACCEPT_DECISIONS:
            final_type = _clean(item.get("revised_entity_type")) or (row and _field(row, "entity_type")) or ""
            final_text = _clean(item.get("revised_entity_text")) or (row and _field(row, "entity_text")) or ""
            evidence = _clean(item.get("evidence_text")) or (row and _field(row, "evidence_text")) or ""
            start = _clean(item.get("evidence_start_char")) or (row and row.get("evidence_start_char")) or ""
            end = _clean(item.get("evidence_end_char")) or (row and row.get("evidence_end_char")) or ""
            if final_type not in ALLOWED_VALUES["entity_type"]:
                errors.append(f"invalid revised entity type for {_clean(item.get('entity_gold_id'))}: {final_type}")
            if not final_text:
                errors.append(f"accepted entity has empty text: {_clean(item.get('entity_gold_id'))}")
            if not _evidence_supported(chunk, evidence, start, end):
                errors.append(f"accepted entity evidence is not supported: {_clean(item.get('entity_gold_id'))}")

    for item in additions:
        if not isinstance(item, dict):
            errors.append("entity addition item is not an object")
            continue
        if _clean(item.get("entity_type")) not in ALLOWED_VALUES["entity_type"]:
            errors.append(f"invalid added entity type: {_clean(item.get('entity_type'))}")
        if not _clean(item.get("entity_text")):
            errors.append("added entity has empty entity_text")
        if not _evidence_supported(chunk, item.get("evidence_text"), item.get("evidence_start_char"), item.get("evidence_end_char")):
            errors.append(f"added entity evidence is not supported: {_clean(item.get('entity_text'))}")
    return errors


def _validate_relationship_response(
    response: dict[str, Any],
    chunk: dict[str, Any],
    existing_rows: list[dict[str, Any]],
    accepted_entities: list[dict[str, Any]],
) -> list[str]:
    errors: list[str] = []
    reviews = response.get("relationship_reviews")
    additions = response.get("relationship_additions")
    if not isinstance(reviews, list):
        return ["relationship_reviews must be an array"]
    if not isinstance(additions, list):
        return ["relationship_additions must be an array"]

    expected_ids = {_field(row, "relationship_gold_id") for row in existing_rows}
    observed_ids = {_clean(item.get("relationship_gold_id")) for item in reviews if isinstance(item, dict)}
    if observed_ids != expected_ids:
        missing = sorted(expected_ids - observed_ids)
        extra = sorted(observed_ids - expected_ids)
        errors.append(f"relationship_reviews must cover exactly the chunk relationship IDs; missing={missing}; extra={extra}")

    existing_by_id = {_field(row, "relationship_gold_id"): row for row in existing_rows}
    endpoint_keys = _accepted_entity_endpoint_keys(accepted_entities)
    for item in reviews:
        if not isinstance(item, dict):
            errors.append("relationship review item is not an object")
            continue
        row = existing_by_id.get(_clean(item.get("relationship_gold_id")))
        decision = _clean(item.get("decision"))
        if decision not in RELATIONSHIP_ACCEPT_DECISIONS | RELATIONSHIP_REJECT_DECISIONS:
            errors.append(f"invalid relationship decision: {decision}")
        if decision in RELATIONSHIP_ACCEPT_DECISIONS:
            final_type = _clean(item.get("relationship_type")) or (row and _field(row, "relationship_type")) or ""
            evidence = _clean(item.get("evidence_text")) or (row and _field(row, "evidence_text")) or ""
            start = _clean(item.get("evidence_start_char")) or (row and row.get("evidence_start_char")) or ""
            end = _clean(item.get("evidence_end_char")) or (row and row.get("evidence_end_char")) or ""
            if final_type not in ALLOWED_VALUES["relationship_type"]:
                errors.append(f"invalid relationship type for {_clean(item.get('relationship_gold_id'))}: {final_type}")
            if not _evidence_supported(chunk, evidence, start, end):
                errors.append(f"accepted relationship evidence is not supported: {_clean(item.get('relationship_gold_id'))}")
            if not _endpoint_matches(
                endpoint_keys,
                item.get("source_entity_id") or (row and row.get("source_entity_id")),
                item.get("source_entity_type") or (row and row.get("source_entity_type")),
                item.get("source_entity_text") or (row and row.get("source_entity_text")),
                item.get("source_normalized_name") or (row and row.get("source_normalized_name")),
            ):
                errors.append(f"accepted relationship source endpoint is not accepted: {_clean(item.get('relationship_gold_id'))}")
            if not _endpoint_matches(
                endpoint_keys,
                item.get("target_entity_id") or (row and row.get("target_entity_id")),
                item.get("target_entity_type") or (row and row.get("target_entity_type")),
                item.get("target_entity_text") or (row and row.get("target_entity_text")),
                item.get("target_normalized_name") or (row and row.get("target_normalized_name")),
            ):
                errors.append(f"accepted relationship target endpoint is not accepted: {_clean(item.get('relationship_gold_id'))}")

    for item in additions:
        if not isinstance(item, dict):
            errors.append("relationship addition item is not an object")
            continue
        if _clean(item.get("relationship_type")) not in ALLOWED_VALUES["relationship_type"]:
            errors.append(f"invalid added relationship type: {_clean(item.get('relationship_type'))}")
        if not _evidence_supported(chunk, item.get("evidence_text"), item.get("evidence_start_char"), item.get("evidence_end_char")):
            errors.append(f"added relationship evidence is not supported: {_clean(item.get('relationship_type'))}")
        if not _endpoint_matches(
            endpoint_keys,
            item.get("source_entity_id"),
            item.get("source_entity_type"),
            item.get("source_entity_text"),
            item.get("source_normalized_name"),
        ):
            errors.append(f"added relationship source endpoint is not accepted: {_clean(item.get('relationship_type'))}")
        if not _endpoint_matches(
            endpoint_keys,
            item.get("target_entity_id"),
            item.get("target_entity_type"),
            item.get("target_entity_text"),
            item.get("target_normalized_name"),
        ):
            errors.append(f"added relationship target endpoint is not accepted: {_clean(item.get('relationship_type'))}")
    return errors


def _accepted_entity_endpoint_keys(rows: list[dict[str, Any]]) -> dict[str, set[str]]:
    keys = {"ids": set(), "text": set(), "normalized": set()}
    for row in rows:
        if _field(row, "annotation_status") != "accepted":
            continue
        entity_id = _field(row, "entity_id")
        entity_type = _field(row, "entity_type").casefold()
        entity_text = _field(row, "entity_text").casefold()
        normalized_name = (_field(row, "normalized_name") or _field(row, "entity_text")).casefold()
        if entity_id:
            keys["ids"].add(entity_id)
        if entity_type and entity_text:
            keys["text"].add(f"{entity_type}|{entity_text}")
        if entity_type and normalized_name:
            keys["normalized"].add(f"{entity_type}|{normalized_name}")
    return keys


def _endpoint_matches(
    endpoint_keys: dict[str, set[str]],
    entity_id: Any,
    entity_type: Any,
    entity_text: Any,
    normalized_name: Any,
) -> bool:
    clean_id = _clean(entity_id)
    if clean_id and clean_id in endpoint_keys["ids"]:
        return True
    clean_type = _clean(entity_type).casefold()
    clean_text = _clean(entity_text).casefold()
    clean_normalized = (_clean(normalized_name) or _clean(entity_text)).casefold()
    return (
        f"{clean_type}|{clean_text}" in endpoint_keys["text"]
        or f"{clean_type}|{clean_normalized}" in endpoint_keys["normalized"]
    )


def _note(decision: str, reason: Any) -> str:
    return f"llm_adjudication={decision}; reason={_clean(reason)}"


def _apply_entity_reviews(
    ws: Any,
    header_map: dict[str, int],
    chunk: dict[str, Any],
    rows_by_id: dict[str, dict[str, Any]],
    response: dict[str, Any],
    annotator: str,
) -> None:
    for item in response["entity_reviews"]:
        row = rows_by_id[_clean(item.get("entity_gold_id"))]
        row_number = int(row["_row_number"])
        decision = _clean(item.get("decision"))
        status = "accepted" if decision in ENTITY_ACCEPT_DECISIONS else "rejected"
        _cell_set(ws, header_map, row_number, "annotation_status", status)
        _cell_set(ws, header_map, row_number, "annotator", annotator)
        _set_if_present(ws, header_map, row_number, "entity_type", item.get("revised_entity_type"))
        _set_if_present(ws, header_map, row_number, "entity_text", item.get("revised_entity_text"))
        _set_if_present(ws, header_map, row_number, "normalized_name", item.get("revised_normalized_name"))
        _set_if_present(ws, header_map, row_number, "evidence_text", item.get("evidence_text"))
        _set_if_present(ws, header_map, row_number, "evidence_start_char", item.get("evidence_start_char"))
        _set_if_present(ws, header_map, row_number, "evidence_end_char", item.get("evidence_end_char"))
        _cell_set(ws, header_map, row_number, "notes", _note(decision, item.get("reason")))

    for item in response["entity_additions"]:
        entity_text = _clean(item.get("entity_text"))
        entity_type = _clean(item.get("entity_type"))
        entity_gold_id = _clean(item.get("entity_gold_id")) or _hash_id(
            "llm_entity",
            _field(chunk, "document_id"),
            _field(chunk, "chunk_id"),
            entity_type,
            entity_text,
        )
        _append_row(
            ws,
            ENTITY_HEADERS,
            {
                "entity_gold_id": entity_gold_id,
                "document_id": _field(chunk, "document_id"),
                "chunk_id": _field(chunk, "chunk_id"),
                "pmid": _field(chunk, "pmid"),
                "pmcid": _field(chunk, "pmcid"),
                "entity_type": entity_type,
                "entity_text": entity_text,
                "normalized_name": _clean(item.get("normalized_name")) or entity_text,
                "evidence_text": _clean(item.get("evidence_text")),
                "evidence_section": _field(chunk, "chunk_section"),
                "evidence_start_char": _clean(item.get("evidence_start_char")),
                "evidence_end_char": _clean(item.get("evidence_end_char")),
                "annotator": annotator,
                "annotation_status": "accepted",
                "notes": _note("added", item.get("reason")),
            },
        )


def _apply_relationship_reviews(
    ws: Any,
    header_map: dict[str, int],
    chunk: dict[str, Any],
    rows_by_id: dict[str, dict[str, Any]],
    response: dict[str, Any],
    annotator: str,
) -> None:
    for item in response["relationship_reviews"]:
        row = rows_by_id[_clean(item.get("relationship_gold_id"))]
        row_number = int(row["_row_number"])
        decision = _clean(item.get("decision"))
        accepted = decision in RELATIONSHIP_ACCEPT_DECISIONS
        _cell_set(ws, header_map, row_number, "annotation_status", "accepted" if accepted else "rejected")
        _cell_set(ws, header_map, row_number, "annotation_decision", "include" if accepted else "exclude")
        _cell_set(ws, header_map, row_number, "annotator", annotator)
        for field in [
            "relationship_type",
            "source_entity_id",
            "source_entity_type",
            "source_entity_text",
            "source_normalized_name",
            "target_entity_id",
            "target_entity_type",
            "target_entity_text",
            "target_normalized_name",
            "evidence_text",
            "evidence_start_char",
            "evidence_end_char",
            "direction_verified",
            "explicit_or_implied",
            "negated",
            "speculative",
        ]:
            _set_if_present(ws, header_map, row_number, field, item.get(field))
        _cell_set(ws, header_map, row_number, "notes", _note(decision, item.get("reason")))

    for item in response["relationship_additions"]:
        rel_type = _clean(item.get("relationship_type"))
        source_text = _clean(item.get("source_entity_text"))
        target_text = _clean(item.get("target_entity_text"))
        relationship_gold_id = _clean(item.get("relationship_gold_id")) or _hash_id(
            "llm_relationship",
            _field(chunk, "document_id"),
            _field(chunk, "chunk_id"),
            rel_type,
            source_text,
            target_text,
        )
        _append_row(
            ws,
            RELATIONSHIP_HEADERS,
            {
                "relationship_gold_id": relationship_gold_id,
                "document_id": _field(chunk, "document_id"),
                "chunk_id": _field(chunk, "chunk_id"),
                "pmid": _field(chunk, "pmid"),
                "pmcid": _field(chunk, "pmcid"),
                "relationship_type": rel_type,
                "source_entity_id": _clean(item.get("source_entity_id")),
                "source_entity_type": _clean(item.get("source_entity_type")),
                "source_entity_text": source_text,
                "source_normalized_name": _clean(item.get("source_normalized_name")) or source_text,
                "target_entity_id": _clean(item.get("target_entity_id")),
                "target_entity_type": _clean(item.get("target_entity_type")),
                "target_entity_text": target_text,
                "target_normalized_name": _clean(item.get("target_normalized_name")) or target_text,
                "evidence_text": _clean(item.get("evidence_text")),
                "evidence_section": _field(chunk, "chunk_section"),
                "evidence_start_char": _clean(item.get("evidence_start_char")),
                "evidence_end_char": _clean(item.get("evidence_end_char")),
                "direction_verified": _clean(item.get("direction_verified")),
                "explicit_or_implied": _clean(item.get("explicit_or_implied")),
                "negated": _clean(item.get("negated")),
                "speculative": _clean(item.get("speculative")),
                "annotation_decision": "include",
                "annotator": annotator,
                "annotation_status": "accepted",
                "notes": _note("added", item.get("reason")),
            },
        )


def _chunk_rows(rows: list[dict[str, Any]], chunk_id: str) -> list[dict[str, Any]]:
    return [row for row in rows if _field(row, "chunk_id") == chunk_id]


def _review_chunk(
    workbook: Any,
    headers: dict[str, dict[str, int]],
    chunk: dict[str, Any],
    review_root: Path,
    language_model: LanguageModel,
    annotator: str,
) -> ChunkReviewResult:
    chunk_id = _field(chunk, "chunk_id")
    pmcid = _field(chunk, "pmcid")
    entity_ws = workbook["gold_entities"]
    relationship_ws = workbook["gold_relationships"]
    chunk_ws = workbook["chunks"]
    all_entities = _rows(entity_ws, headers["gold_entities"])
    all_relationships = _rows(relationship_ws, headers["gold_relationships"])
    chunk_entities = _chunk_rows(all_entities, chunk_id)
    chunk_relationships = _chunk_rows(all_relationships, chunk_id)

    entity_response, entity_audit_path, entity_error = _call_json_model(
        language_model,
        _entity_prompt(chunk, chunk_entities),
        _entity_review_schema(),
        _model_call_path(review_root, chunk, "entity_adjudication"),
    )
    if entity_error:
        return ChunkReviewResult(chunk_id, pmcid, "error", entity_audit_path, "", entity_error)
    entity_errors = _validate_entity_response(entity_response, chunk, chunk_entities)
    if entity_errors:
        return ChunkReviewResult(chunk_id, pmcid, "error", entity_audit_path, "", "; ".join(entity_errors))

    _apply_entity_reviews(
        entity_ws,
        headers["gold_entities"],
        chunk,
        {_field(row, "entity_gold_id"): row for row in chunk_entities},
        entity_response,
        annotator,
    )

    refreshed_entities = _chunk_rows(_rows(entity_ws, headers["gold_entities"]), chunk_id)
    relationship_response, relationship_audit_path, relationship_error = _call_json_model(
        language_model,
        _relationship_prompt(chunk, refreshed_entities, chunk_relationships),
        _relationship_review_schema(),
        _model_call_path(review_root, chunk, "relationship_adjudication"),
    )
    if relationship_error:
        return ChunkReviewResult(chunk_id, pmcid, "error", entity_audit_path, relationship_audit_path, relationship_error)
    relationship_errors = _validate_relationship_response(relationship_response, chunk, chunk_relationships, refreshed_entities)
    if relationship_errors:
        return ChunkReviewResult(
            chunk_id,
            pmcid,
            "error",
            entity_audit_path,
            relationship_audit_path,
            "; ".join(relationship_errors),
        )

    _apply_relationship_reviews(
        relationship_ws,
        headers["gold_relationships"],
        chunk,
        {_field(row, "relationship_gold_id"): row for row in chunk_relationships},
        relationship_response,
        annotator,
    )
    _cell_set(chunk_ws, headers["chunks"], int(chunk["_row_number"]), "annotation_status", "reviewed")
    _cell_set(chunk_ws, headers["chunks"], int(chunk["_row_number"]), "annotator", annotator)
    _cell_set(chunk_ws, headers["chunks"], int(chunk["_row_number"]), "notes", "llm_adjudication=reviewed")
    return ChunkReviewResult(chunk_id, pmcid, "ok", entity_audit_path, relationship_audit_path)


def run_llm_adjudication(
    workbook_path: Path,
    review_root: Path,
    *,
    model_profile: str = "frontier",
    model: str | None = None,
) -> LlmAdjudicationResult:
    profile_dict, language_model = _frontier_language_model(model_profile, model)
    annotator = f"{profile_dict['name']} / {language_model.provider} / {language_model.model}"
    workbook = load_workbook(workbook_path)
    try:
        headers = {name: _headers(workbook[name]) for name in ["chunks", "gold_entities", "gold_relationships"]}
        chunks = _rows(workbook["chunks"], headers["chunks"])
        results: list[ChunkReviewResult] = []
        for chunk in chunks:
            results.append(_review_chunk(workbook, headers, chunk, review_root, language_model, annotator))
        workbook.save(workbook_path)
    finally:
        workbook.close()

    audit_paths: list[str] = []
    for result in results:
        if result.entity_audit_path:
            audit_paths.append(result.entity_audit_path)
        if result.relationship_audit_path:
            audit_paths.append(result.relationship_audit_path)
    return LlmAdjudicationResult(
        llm_review=True,
        model_profile=profile_dict,
        provider=language_model.provider,
        model=language_model.model,
        reviewed_chunk_count=sum(1 for result in results if result.status == "ok"),
        failed_chunk_count=sum(1 for result in results if result.status != "ok"),
        model_call_paths=audit_paths,
        chunks=results,
    )


__all__ = [
    "ChunkReviewResult",
    "LlmAdjudicationResult",
    "PROMPT_VERSION",
    "run_llm_adjudication",
]
