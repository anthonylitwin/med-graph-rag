from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


DOCUMENT_HEADERS = [
    "document_id",
    "pmid",
    "pmcid",
    "title",
    "year",
    "journal",
    "doi",
    "authors",
    "abstract",
    "topic_tags",
    "included_in_extraction_gold",
    "included_in_qa_eval",
    "source_file_path",
    "source_url",
    "download_date",
    "notes",
]
CHUNK_HEADERS = [
    "chunk_id",
    "document_id",
    "pmid",
    "pmcid",
    "chunk_index",
    "chunk_section",
    "chunk_text",
    "start_char",
    "end_char",
    "included_in_gold_annotation",
    "annotator",
    "annotation_status",
    "notes",
]
ENTITY_HEADERS = [
    "entity_gold_id",
    "document_id",
    "chunk_id",
    "pmid",
    "pmcid",
    "entity_id",
    "entity_type",
    "entity_text",
    "normalized_name",
    "aliases",
    "category",
    "mesh_id",
    "icd10",
    "drugbank_id",
    "unit",
    "normal_range",
    "description",
    "evidence_text",
    "evidence_section",
    "evidence_start_char",
    "evidence_end_char",
    "annotator",
    "annotation_status",
    "notes",
]
RELATIONSHIP_HEADERS = [
    "relationship_gold_id",
    "document_id",
    "chunk_id",
    "pmid",
    "pmcid",
    "relationship_type",
    "source_entity_id",
    "source_entity_type",
    "source_entity_text",
    "source_normalized_name",
    "target_entity_id",
    "target_entity_type",
    "target_entity_text",
    "target_normalized_name",
    "evidence_text",
    "evidence_section",
    "evidence_start_char",
    "evidence_end_char",
    "confidence_gold",
    "direction_verified",
    "explicit_or_implied",
    "negated",
    "speculative",
    "annotation_decision",
    "annotator",
    "annotation_status",
    "notes",
]
REJECTED_HEADERS = [
    "rejected_id",
    "document_id",
    "chunk_id",
    "pmid",
    "pmcid",
    "candidate_text",
    "candidate_type",
    "candidate_relationship_type",
    "source_candidate",
    "target_candidate",
    "reason",
    "evidence_text",
    "annotation_rule",
    "annotator",
    "notes",
]
NOTE_HEADERS = [
    "note_id",
    "document_id",
    "chunk_id",
    "related_entity_gold_id",
    "related_relationship_gold_id",
    "issue_type",
    "question",
    "decision",
    "rationale",
    "annotator",
    "created_at",
]
ALLOWED_VALUE_HEADERS = [
    "entity_type",
    "relationship_type",
    "annotation_status",
    "annotation_decision",
    "yes_no_unclear",
    "direction_verified",
    "explicit_or_implied",
    "candidate_type",
    "issue_type",
    "boolean_yes_no",
]

ALLOWED_VALUES = {
    "entity_type": ["Drug", "Condition", "Symptom", "RiskFactor", "Biomarker", "Paper"],
    "relationship_type": [
        "TREATS",
        "PREVENTS",
        "REDUCES",
        "INCREASES",
        "ASSOCIATED_WITH",
        "HAS_ADVERSE_EFFECT",
        "CAUSES",
        "HAS_SYMPTOM",
        "INCREASES_RISK_OF",
        "INTERACTS_WITH",
        "CONTRAINDICATED_FOR",
        "MENTIONS",
    ],
    "annotation_status": ["draft", "needs_review", "accepted", "rejected", "reviewed"],
    "annotation_decision": ["include", "exclude", "needs_review"],
    "yes_no_unclear": ["yes", "no", "unclear"],
    "direction_verified": ["yes", "no", "needs_review"],
    "explicit_or_implied": ["explicit", "strong_implication", "weak_implication"],
    "candidate_type": ["entity", "relationship", "other"],
    "issue_type": [
        "alias_normalization",
        "relationship_ambiguity",
        "directionality_issue",
        "ontology_gap",
        "evidence_unclear",
        "duplicate_entity",
        "other",
    ],
    "boolean_yes_no": ["yes", "no"],
}

VALIDATIONS = {
    "documents": [
        ("K2:K5000", "allowed_values!$J$2:$J$3"),
        ("L2:L5000", "allowed_values!$J$2:$J$3"),
    ],
    "chunks": [
        ("J2:J5000", "allowed_values!$J$2:$J$3"),
        ("L2:L5000", "allowed_values!$C$2:$C$6"),
    ],
    "gold_entities": [
        ("G2:G5000", "allowed_values!$A$2:$A$7"),
        ("W2:W5000", "allowed_values!$C$2:$C$6"),
    ],
    "gold_relationships": [
        ("H2:H5000", "allowed_values!$A$2:$A$7"),
        ("L2:L5000", "allowed_values!$A$2:$A$7"),
        ("F2:F5000", "allowed_values!$B$2:$B$13"),
        ("T2:T5000", "allowed_values!$F$2:$F$4"),
        ("U2:U5000", "allowed_values!$G$2:$G$4"),
        ("V2:V5000", "allowed_values!$E$2:$E$4"),
        ("W2:W5000", "allowed_values!$E$2:$E$4"),
        ("X2:X5000", "allowed_values!$D$2:$D$4"),
        ("Z2:Z5000", "allowed_values!$C$2:$C$6"),
    ],
    "rejected_candidates": [
        ("G2:G5000", "allowed_values!$H$2:$H$4"),
        ("H2:H5000", "allowed_values!$B$2:$B$13"),
    ],
    "annotation_notes": [("F2:F5000", "allowed_values!$I$2:$I$8")],
}


def _hash_id(prefix: str, *parts: Any) -> str:
    payload = "|".join(str(part or "") for part in parts)
    return f"{prefix}:{hashlib.sha1(payload.encode('utf-8')).hexdigest()[:16]}"


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=True, sort_keys=True)
    return str(value)


def _record_annotator(record: dict[str, Any]) -> str:
    run = record.get("run") if isinstance(record.get("run"), dict) else {}
    profile = str(run.get("model_profile") or "")
    model = str(run.get("extractor_model") or "")
    provider = str(run.get("extractor_provider") or "")
    return " / ".join(item for item in (profile, provider, model) if item)


def _span(text: str, needle: str, base_offset: int = 0) -> tuple[str, int | str, int | str]:
    if not text or not needle:
        return needle, "", ""
    index = text.lower().find(needle.lower())
    if index < 0:
        return needle, "", ""
    end = index + len(needle)
    return text[index:end], base_offset + index, base_offset + end


def _chunk_by_id(record: dict[str, Any]) -> dict[str, dict[str, Any]]:
    chunks = record.get("chunks") if isinstance(record.get("chunks"), list) else []
    return {str(chunk.get("id")): chunk for chunk in chunks if isinstance(chunk, dict)}


def _sorted_extractions(record: dict[str, Any]) -> list[dict[str, Any]]:
    chunk_order = {
        str(chunk.get("id")): int(chunk.get("order") or 0)
        for chunk in record.get("chunks", [])
        if isinstance(chunk, dict)
    }
    extractions = record.get("extractions") if isinstance(record.get("extractions"), list) else []
    return sorted(
        [item for item in extractions if isinstance(item, dict)],
        key=lambda item: (chunk_order.get(str(item.get("chunk_id")), 0), str(item.get("chunk_id") or "")),
    )


def _candidate_type(reason: str, candidate_text: str) -> str:
    lowered = f"{reason} {candidate_text}".lower()
    if "relationship" in lowered:
        return "relationship"
    if "entity" in lowered or "gliner" in lowered:
        return "entity"
    return "other"


def _audit_note(paths: Any) -> str:
    if not isinstance(paths, list) or not paths:
        return "silver bootstrap"
    return "silver bootstrap; audit=" + ";".join(str(path) for path in paths)


def _append_row(ws, headers: list[str], row: dict[str, Any]) -> None:
    ws.append([row.get(header, "") for header in headers])


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        header = str(column_cells[0].value or "")
        width = min(max(len(header) + 2, 12), 60)
        ws.column_dimensions[column_cells[0].column_letter].width = width


def _add_validations(wb: Workbook) -> None:
    for sheet_name, validations in VALIDATIONS.items():
        ws = wb[sheet_name]
        for cell_range, formula in validations:
            validation = DataValidation(type="list", formula1=formula, allow_blank=True)
            ws.add_data_validation(validation)
            validation.add(cell_range)


def _build_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("README")
    rows = [
        ("MedGraphRAG v1.1 Gold Annotation Workbook",),
        (),
        ("Purpose", "Use this workbook to review silver model annotations before promoting rows to gold labels."),
        (
            "Recommended workflow",
            "Review documents and chunks, then accept, reject, or edit rows in gold_entities and gold_relationships.",
        ),
        ("Silver bootstrap", "Rows generated by annotation bootstrap default to needs_review."),
        ("Required evidence", "Every accepted relationship should keep concise evidence_text copied from the source chunk."),
        ("Directionality", "Relationship direction matters. Example: Drug REDUCES Biomarker."),
        ("Rejected candidates", "Use rejected_candidates for tempting but unsupported labels and prompt-tuning notes."),
        ("Status values", "Use draft, needs_review, accepted, rejected, or reviewed to track annotation progress."),
        (),
        (),
        ("Sheet", "Use", "Key Required Fields", "Notes"),
        ("documents", "Benchmark document metadata", "document_id, pmid/pmcid, title", "One row per source paper."),
        ("chunks", "Chunk provenance", "chunk_id, document_id, chunk_section, chunk_text", "One row per extraction chunk."),
        (
            "gold_entities",
            "Human-review entity labels",
            "entity_gold_id, document_id, chunk_id, entity_id, entity_type, entity_text, evidence_text",
            "Silver suggestions start as needs_review.",
        ),
        (
            "gold_relationships",
            "Human-review relationship triples",
            "relationship_gold_id, source, relationship_type, target, evidence_text",
            "Most important sheet for GraphRAG evaluation.",
        ),
        (
            "rejected_candidates",
            "Unsupported or invalid candidate labels",
            "candidate_text, reason, evidence_text",
            "Useful for error analysis and prompt tuning.",
        ),
        ("annotation_notes", "Reviewer decisions and uncertainty", "issue_type, question, decision, rationale", ""),
        ("allowed_values", "Dropdown source lists", "N/A", "Reference sheet for validation lists."),
    ]
    for row in rows:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 92
    ws.column_dimensions["C"].width = 62
    ws.column_dimensions["D"].width = 60


def _build_allowed_values(wb: Workbook) -> None:
    ws = wb.create_sheet("allowed_values")
    ws.append(ALLOWED_VALUE_HEADERS)
    max_len = max(len(values) for values in ALLOWED_VALUES.values())
    for index in range(max_len):
        ws.append([ALLOWED_VALUES[header][index] if index < len(ALLOWED_VALUES[header]) else "" for header in ALLOWED_VALUE_HEADERS])
    _style_sheet(ws)


def _build_documents_sheet(
    wb: Workbook,
    records: list[dict[str, Any]],
    processed_paths: dict[str, str],
) -> None:
    ws = wb.create_sheet("documents")
    ws.append(DOCUMENT_HEADERS)
    for record in records:
        document = record.get("document") if isinstance(record.get("document"), dict) else {}
        document_id = str(document.get("id") or "")
        pmcid = str(document.get("pmcid") or "")
        _append_row(
            ws,
            DOCUMENT_HEADERS,
            {
                "document_id": document_id,
                "pmid": document.get("pmid", ""),
                "pmcid": pmcid,
                "title": document.get("title", ""),
                "year": document.get("year", ""),
                "journal": document.get("journal", ""),
                "doi": document.get("doi", ""),
                "authors": _as_text(document.get("authors", "")),
                "abstract": document.get("abstract", ""),
                "included_in_extraction_gold": "yes",
                "included_in_qa_eval": "yes",
                "source_file_path": processed_paths.get(document_id) or processed_paths.get(pmcid, ""),
                "source_url": document.get("source_url", ""),
                "download_date": document.get("ingested_at", ""),
                "notes": "silver bootstrap source document",
            },
        )
    _style_sheet(ws)


def _build_chunks_sheet(wb: Workbook, records: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("chunks")
    ws.append(CHUNK_HEADERS)
    for record in records:
        document = record.get("document") if isinstance(record.get("document"), dict) else {}
        annotator = _record_annotator(record)
        chunks = record.get("chunks") if isinstance(record.get("chunks"), list) else []
        for chunk in sorted([item for item in chunks if isinstance(item, dict)], key=lambda item: int(item.get("order") or 0)):
            _append_row(
                ws,
                CHUNK_HEADERS,
                {
                    "chunk_id": chunk.get("id", ""),
                    "document_id": chunk.get("document_id") or document.get("id", ""),
                    "pmid": document.get("pmid", ""),
                    "pmcid": document.get("pmcid", ""),
                    "chunk_index": chunk.get("order", ""),
                    "chunk_section": chunk.get("section", ""),
                    "chunk_text": chunk.get("text", ""),
                    "start_char": chunk.get("char_start", ""),
                    "end_char": chunk.get("char_end", ""),
                    "included_in_gold_annotation": "yes",
                    "annotator": annotator,
                    "annotation_status": "needs_review",
                    "notes": "silver bootstrap chunk",
                },
            )
    _style_sheet(ws)


def _build_entities_sheet(wb: Workbook, records: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("gold_entities")
    ws.append(ENTITY_HEADERS)
    for record in records:
        document = record.get("document") if isinstance(record.get("document"), dict) else {}
        pmcid = str(document.get("pmcid") or "")
        pmid = str(document.get("pmid") or "")
        annotator = _record_annotator(record)
        chunks = _chunk_by_id(record)
        for extraction in _sorted_extractions(record):
            chunk_id = str(extraction.get("chunk_id") or "")
            chunk = chunks.get(chunk_id, {})
            chunk_text = str(chunk.get("text") or "")
            base_offset = int(chunk.get("char_start") or 0)
            entities = extraction.get("entities") if isinstance(extraction.get("entities"), list) else []
            for entity in sorted([item for item in entities if isinstance(item, dict)], key=lambda item: (str(item.get("type")), str(item.get("name")), str(item.get("id")))):
                props = entity.get("properties") if isinstance(entity.get("properties"), dict) else {}
                entity_text = str(entity.get("name") or "")
                evidence_text, start_char, end_char = _span(chunk_text, entity_text, base_offset)
                _append_row(
                    ws,
                    ENTITY_HEADERS,
                    {
                        "entity_gold_id": _hash_id("silver_entity", document.get("id"), chunk_id, entity.get("id"), entity_text),
                        "document_id": document.get("id", ""),
                        "chunk_id": chunk_id,
                        "pmid": pmid,
                        "pmcid": pmcid,
                        "entity_id": entity.get("id", ""),
                        "entity_type": entity.get("type", ""),
                        "entity_text": entity_text,
                        "normalized_name": entity_text,
                        "aliases": _as_text(props.get("aliases", "")),
                        "category": props.get("category", ""),
                        "mesh_id": props.get("mesh_id", ""),
                        "icd10": props.get("icd10", ""),
                        "drugbank_id": props.get("drugbank_id", ""),
                        "unit": props.get("unit", ""),
                        "normal_range": props.get("normal_range", ""),
                        "description": props.get("description", ""),
                        "evidence_text": evidence_text,
                        "evidence_section": chunk.get("section", ""),
                        "evidence_start_char": start_char,
                        "evidence_end_char": end_char,
                        "annotator": annotator,
                        "annotation_status": "needs_review",
                        "notes": _audit_note(extraction.get("model_call_paths")),
                    },
                )
    _style_sheet(ws)


def _build_relationships_sheet(wb: Workbook, records: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("gold_relationships")
    ws.append(RELATIONSHIP_HEADERS)
    for record in records:
        document = record.get("document") if isinstance(record.get("document"), dict) else {}
        pmcid = str(document.get("pmcid") or "")
        pmid = str(document.get("pmid") or "")
        annotator = _record_annotator(record)
        chunks = _chunk_by_id(record)
        for extraction in _sorted_extractions(record):
            chunk_id = str(extraction.get("chunk_id") or "")
            chunk = chunks.get(chunk_id, {})
            chunk_text = str(chunk.get("text") or "")
            base_offset = int(chunk.get("char_start") or 0)
            relationships = extraction.get("relationships") if isinstance(extraction.get("relationships"), list) else []
            for relationship in sorted([item for item in relationships if isinstance(item, dict)], key=lambda item: str(item.get("id"))):
                props = relationship.get("properties") if isinstance(relationship.get("properties"), dict) else {}
                source = relationship.get("source") if isinstance(relationship.get("source"), dict) else {}
                target = relationship.get("target") if isinstance(relationship.get("target"), dict) else {}
                evidence = str(props.get("evidence") or "")
                evidence_text, start_char, end_char = _span(chunk_text, evidence, base_offset)
                _append_row(
                    ws,
                    RELATIONSHIP_HEADERS,
                    {
                        "relationship_gold_id": _hash_id(
                            "silver_relationship",
                            document.get("id"),
                            chunk_id,
                            relationship.get("id"),
                            relationship.get("type"),
                            source.get("id"),
                            target.get("id"),
                        ),
                        "document_id": document.get("id", ""),
                        "chunk_id": chunk_id,
                        "pmid": pmid,
                        "pmcid": pmcid,
                        "relationship_type": relationship.get("type", ""),
                        "source_entity_id": source.get("id", ""),
                        "source_entity_type": source.get("type", ""),
                        "source_entity_text": source.get("name", ""),
                        "source_normalized_name": source.get("name", ""),
                        "target_entity_id": target.get("id", ""),
                        "target_entity_type": target.get("type", ""),
                        "target_entity_text": target.get("name", ""),
                        "target_normalized_name": target.get("name", ""),
                        "evidence_text": evidence_text,
                        "evidence_section": chunk.get("section", ""),
                        "evidence_start_char": start_char,
                        "evidence_end_char": end_char,
                        "confidence_gold": props.get("confidence", ""),
                        "direction_verified": "needs_review",
                        "explicit_or_implied": "explicit" if evidence else "",
                        "negated": "unclear",
                        "speculative": "unclear",
                        "annotation_decision": "needs_review",
                        "annotator": annotator,
                        "annotation_status": "needs_review",
                        "notes": _audit_note(extraction.get("model_call_paths")),
                    },
                )
    _style_sheet(ws)


def _build_rejected_sheet(wb: Workbook, records: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("rejected_candidates")
    ws.append(REJECTED_HEADERS)
    for record in records:
        document = record.get("document") if isinstance(record.get("document"), dict) else {}
        pmcid = str(document.get("pmcid") or "")
        pmid = str(document.get("pmid") or "")
        annotator = _record_annotator(record)
        for extraction in _sorted_extractions(record):
            chunk_id = str(extraction.get("chunk_id") or "")
            rejected = extraction.get("rejected_candidates") if isinstance(extraction.get("rejected_candidates"), list) else []
            for item in rejected:
                if not isinstance(item, dict):
                    candidate_text = str(item)
                    reason = "rejected candidate is not an object"
                else:
                    candidate_text = str(item.get("text") or "")
                    reason = str(item.get("reason") or "")
                _append_row(
                    ws,
                    REJECTED_HEADERS,
                    {
                        "rejected_id": _hash_id("silver_rejected", document.get("id"), chunk_id, candidate_text, reason),
                        "document_id": document.get("id", ""),
                        "chunk_id": chunk_id,
                        "pmid": pmid,
                        "pmcid": pmcid,
                        "candidate_text": candidate_text,
                        "candidate_type": _candidate_type(reason, candidate_text),
                        "reason": reason,
                        "annotation_rule": reason,
                        "annotator": annotator,
                        "notes": _audit_note(extraction.get("model_call_paths")),
                    },
                )
    _style_sheet(ws)


def _build_notes_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("annotation_notes")
    ws.append(NOTE_HEADERS)
    _style_sheet(ws)


def export_annotation_workbook(
    processed_records: list[dict[str, Any]],
    workbook_path: Path,
    *,
    processed_paths: dict[str, str] | None = None,
) -> Path:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    records = sorted(
        processed_records,
        key=lambda record: str((record.get("document") if isinstance(record.get("document"), dict) else {}).get("pmcid", "")),
    )
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    _build_readme(wb)
    _build_documents_sheet(wb, records, processed_paths or {})
    _build_chunks_sheet(wb, records)
    _build_entities_sheet(wb, records)
    _build_relationships_sheet(wb, records)
    _build_rejected_sheet(wb, records)
    _build_notes_sheet(wb)
    _build_allowed_values(wb)
    _add_validations(wb)

    wb.save(workbook_path)
    return workbook_path


__all__ = ["export_annotation_workbook"]
