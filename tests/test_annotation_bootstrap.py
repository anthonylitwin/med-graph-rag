from __future__ import annotations

import json
import tempfile
import types
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook

from pipelines.annotation.bootstrap_annotations import parse_args, run_bootstrap
from pipelines.annotation.workbook import export_annotation_workbook


def _fixture_processed_record() -> dict:
    return {
        "run": {
            "id": "fixture-run",
            "created_at": "2026-01-01T00:00:00+00:00",
            "source": "pmc_bioc",
            "model_profile": "noop",
            "extractor_provider": "fixture",
            "extractor_model": "fixture-model",
            "prompt_version": "001_initial_prompt",
            "min_confidence": 0.5,
        },
        "document": {
            "id": "paper:PMC123",
            "pmcid": "PMC123",
            "pmid": "123",
            "title": "Fish oil fixture",
            "year": "2026",
            "journal": "Fixture Journal",
            "doi": "10.0000/fixture",
            "authors": ["A Reviewer", "B Curator"],
            "source_url": "https://pmc.ncbi.nlm.nih.gov/articles/PMC123/",
            "ingested_at": "2026-01-01T00:00:00+00:00",
        },
        "chunks": [
            {
                "id": "PMC123-chunk-0001",
                "document_id": "paper:PMC123",
                "pmcid": "PMC123",
                "order": 1,
                "char_start": 100,
                "char_end": 142,
                "section": "Abstract",
                "type": "abstract",
                "source_sections": ["Abstract"],
                "text": "Fish oil reduced triglycerides in adults.",
            }
        ],
        "extractions": [
            {
                "chunk_id": "PMC123-chunk-0001",
                "status": "ok",
                "entities": [
                    {
                        "id": "drug:fish_oil",
                        "type": "Drug",
                        "name": "Fish oil",
                        "properties": {"source": "pmc", "extractor": "fixture", "model": "fixture-model"},
                    },
                    {
                        "id": "biomarker:triglycerides",
                        "type": "Biomarker",
                        "name": "Triglycerides",
                        "properties": {"source": "pmc", "extractor": "fixture", "model": "fixture-model"},
                    },
                ],
                "relationships": [
                    {
                        "id": "rel:abc",
                        "type": "REDUCES",
                        "source": {"id": "drug:fish_oil", "type": "Drug", "name": "Fish oil"},
                        "target": {"id": "biomarker:triglycerides", "type": "Biomarker", "name": "Triglycerides"},
                        "properties": {
                            "confidence": 0.91,
                            "evidence": "Fish oil reduced triglycerides",
                            "source_pmcid": "PMC123",
                            "chunk_id": "PMC123-chunk-0001",
                        },
                    }
                ],
                "rejected_candidates": [{"text": "adults", "reason": "entity has unsupported type or missing name"}],
                "model_call_paths": ["model_calls/PMC123/PMC123-chunk-0001.extraction.json"],
            }
        ],
        "entities": [],
        "relationships": [],
        "rejected_candidates": [],
    }


def _row_dict(ws, row_number: int) -> dict:
    headers = [cell.value for cell in ws[1]]
    values = [cell.value for cell in ws[row_number]]
    return dict(zip(headers, values))


class AnnotationWorkbookTests(unittest.TestCase):
    def test_export_workbook_uses_template_sheets_review_statuses_and_stable_ids(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            first_path = Path(tmpdir) / "first.xlsx"
            second_path = Path(tmpdir) / "second.xlsx"
            record = _fixture_processed_record()
            export_annotation_workbook([record], first_path, processed_paths={"paper:PMC123": "processed/PMC123.json"})
            export_annotation_workbook([record], second_path, processed_paths={"paper:PMC123": "processed/PMC123.json"})

            workbook = load_workbook(first_path)
            second = load_workbook(second_path)

        self.assertEqual(
            workbook.sheetnames,
            [
                "README",
                "documents",
                "chunks",
                "gold_entities",
                "gold_relationships",
                "rejected_candidates",
                "annotation_notes",
                "allowed_values",
            ],
        )
        document = _row_dict(workbook["documents"], 2)
        chunk = _row_dict(workbook["chunks"], 2)
        entities = [_row_dict(workbook["gold_entities"], row_number) for row_number in range(2, 4)]
        entity = next(row for row in entities if row["entity_text"] == "Fish oil")
        relationship = _row_dict(workbook["gold_relationships"], 2)
        rejected = _row_dict(workbook["rejected_candidates"], 2)
        second_entities = [_row_dict(second["gold_entities"], row_number) for row_number in range(2, 4)]
        second_entity = next(row for row in second_entities if row["entity_text"] == "Fish oil")

        self.assertEqual(document["included_in_extraction_gold"], "yes")
        self.assertEqual(document["source_file_path"], "processed/PMC123.json")
        self.assertEqual(chunk["annotation_status"], "needs_review")
        self.assertEqual(entity["annotation_status"], "needs_review")
        self.assertEqual(entity["evidence_start_char"], 100)
        self.assertEqual(entity["entity_gold_id"], second_entity["entity_gold_id"])
        self.assertEqual(relationship["annotation_decision"], "needs_review")
        self.assertEqual(relationship["direction_verified"], "needs_review")
        self.assertEqual(relationship["negated"], "unclear")
        self.assertEqual(relationship["confidence_gold"], 0.91)
        self.assertEqual(rejected["candidate_type"], "entity")
        self.assertGreater(len(workbook["gold_relationships"].data_validations.dataValidation), 0)


class AnnotationBootstrapCliTests(unittest.TestCase):
    def test_run_bootstrap_orchestrates_ingestion_and_writes_manifest_and_workbook(self) -> None:
        record = _fixture_processed_record()

        def fake_process(config):
            processed_dir = config.output_root / "processed"
            processed_dir.mkdir(parents=True, exist_ok=True)
            processed_path = processed_dir / "PMC123.json"
            processed_path.write_text(json.dumps(record), encoding="utf-8")
            return [
                types.SimpleNamespace(
                    pmcid="PMC123",
                    status="ok",
                    processed_path=processed_path,
                    manifest_row=lambda: {
                        "pmcid": "PMC123",
                        "processed_path": processed_path.as_posix(),
                        "status": "ok",
                    },
                )
            ]

        with tempfile.TemporaryDirectory() as tmpdir:
            args = parse_args(["--pmcid", "PMC123", "--output-root", tmpdir, "--model-profile", "noop"])
            with mock.patch("pipelines.annotation.bootstrap_annotations.process_pmc_articles", fake_process):
                manifest = run_bootstrap(args)

            workbook_path = Path(manifest["workbook_path"])
            run_manifest_path = Path(manifest["run_manifest_path"])
            top_manifest_path = Path(manifest["manifest_path"])

            self.assertTrue(workbook_path.exists())
            self.assertTrue(run_manifest_path.exists())
            self.assertTrue(top_manifest_path.exists())
            self.assertEqual(manifest["success_count"], 1)
            self.assertIn("model_calls", manifest["model_calls_root"])
            self.assertEqual(load_workbook(workbook_path)["documents"]["A2"].value, "paper:PMC123")


if __name__ == "__main__":
    unittest.main()
