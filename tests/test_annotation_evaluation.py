from __future__ import annotations

import json
import sys
import tempfile
import types
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook

from pipelines.annotation.evaluate_annotations import parse_args
from pipelines.annotation.evaluation import AnnotationEvaluationConfig, run_annotation_evaluation
from pipelines.annotation.gold_export import write_gold_exports
from pipelines.annotation.review_workbook import read_review_workbook
from pipelines.annotation.workbook import export_annotation_workbook
from pipelines.ingestion.extractors import StaticFixtureExtractor


def _fixture_processed_record() -> dict:
    return {
        "run": {
            "id": "fixture-run",
            "created_at": "2026-01-01T00:00:00+00:00",
            "source": "pmc_bioc",
            "model_profile": "fixture",
            "extractor_provider": "fixture",
            "extractor_model": "fixture-model",
            "prompt_version": "001_initial_prompt",
            "min_confidence": 0.5,
        },
        "document": {
            "id": "paper:PMC123",
            "pmcid": "PMC123",
            "pmid": "123",
            "title": "Fish oil fixture",
            "year": "2026",
            "journal": "Fixture Journal",
            "doi": "10.0000/fixture",
            "authors": ["A Reviewer"],
            "source_url": "https://pmc.ncbi.nlm.nih.gov/articles/PMC123/",
        },
        "chunks": [
            {
                "id": "PMC123-chunk-0001",
                "document_id": "paper:PMC123",
                "pmcid": "PMC123",
                "order": 1,
                "char_start": 100,
                "char_end": 142,
                "section": "Abstract",
                "type": "abstract",
                "source_sections": ["Abstract"],
                "text": "Fish oil reduced triglycerides in adults.",
            }
        ],
        "extractions": [
            {
                "chunk_id": "PMC123-chunk-0001",
                "status": "ok",
                "entities": [
                    {"id": "drug:fish_oil", "type": "Drug", "name": "Fish oil", "properties": {}},
                    {"id": "biomarker:triglycerides", "type": "Biomarker", "name": "Triglycerides", "properties": {}},
                ],
                "relationships": [
                    {
                        "id": "rel:abc",
                        "type": "REDUCES",
                        "source": {"id": "drug:fish_oil", "type": "Drug", "name": "Fish oil"},
                        "target": {"id": "biomarker:triglycerides", "type": "Biomarker", "name": "Triglycerides"},
                        "properties": {
                            "confidence": 0.91,
                            "evidence": "Fish oil reduced triglycerides",
                        },
                    }
                ],
                "rejected_candidates": [],
            }
        ],
    }


def _column_index(ws, header: str) -> int:
    headers = [cell.value for cell in ws[1]]
    return headers.index(header) + 1


def _set_gold_statuses(workbook_path: Path) -> None:
    workbook = load_workbook(workbook_path)
    chunks = workbook["chunks"]
    chunk_status_col = _column_index(chunks, "annotation_status")
    for row in range(2, chunks.max_row + 1):
        chunks.cell(row=row, column=chunk_status_col).value = "reviewed"

    entities = workbook["gold_entities"]
    entity_status_col = _column_index(entities, "annotation_status")
    for row in range(2, entities.max_row + 1):
        entities.cell(row=row, column=entity_status_col).value = "accepted"

    relationships = workbook["gold_relationships"]
    relationship_status_col = _column_index(relationships, "annotation_status")
    decision_col = _column_index(relationships, "annotation_decision")
    direction_col = _column_index(relationships, "direction_verified")
    negated_col = _column_index(relationships, "negated")
    speculative_col = _column_index(relationships, "speculative")
    for row in range(2, relationships.max_row + 1):
        relationships.cell(row=row, column=relationship_status_col).value = "accepted"
        relationships.cell(row=row, column=decision_col).value = "include"
        relationships.cell(row=row, column=direction_col).value = "yes"
        relationships.cell(row=row, column=negated_col).value = "no"
        relationships.cell(row=row, column=speculative_col).value = "no"
    workbook.save(workbook_path)


def _write_gold_manifest(root: Path) -> Path:
    workbook_path = root / "reviewed_annotation_workbook.xlsx"
    export_annotation_workbook([_fixture_processed_record()], workbook_path)
    _set_gold_statuses(workbook_path)
    write_gold_exports(read_review_workbook(workbook_path), root)
    manifest = {
        "schema_version": "gold_manifest_v1",
        "gold_set_id": "fixture_gold",
        "artifacts": {
            "reviewed_workbook": {"path": workbook_path.as_posix()},
            "gold_entities": {"path": (root / "gold_entities.csv").as_posix()},
            "gold_relationships": {"path": (root / "gold_relationships.csv").as_posix()},
        },
    }
    manifest_path = root / "gold_manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return manifest_path


class AnnotationEvaluationTests(unittest.TestCase):
    def test_parse_args_accepts_gold_manifest_and_eval_controls(self) -> None:
        args = parse_args(
            [
                "--gold-manifest",
                "gold_manifest.json",
                "--eval-id",
                "eval-fixture",
                "--model-profile",
                "noop",
                "--limit",
                "1",
                "--force",
            ]
        )

        self.assertEqual(args.gold_manifest, Path("gold_manifest.json"))
        self.assertEqual(args.eval_id, "eval-fixture")
        self.assertEqual(args.model_profile, "noop")
        self.assertEqual(args.entity_threshold, 0.5)
        self.assertEqual(args.concept_threshold, 0.84)
        self.assertEqual(args.relation_threshold, 0.66)
        self.assertEqual(args.limit, 1)
        self.assertEqual(args.neo4j_load_mode, "none")
        self.assertFalse(args.apply_schema)
        self.assertEqual(args.neo4j_run_label, "")
        self.assertFalse(args.mlflow)
        self.assertEqual(args.mlflow_tracking_uri, "http://localhost:5000")
        self.assertEqual(args.mlflow_experiment, "medgraphrag-annotation-eval")
        self.assertEqual(args.mlflow_run_name, "")
        self.assertFalse(args.no_mlflow_artifacts)
        self.assertTrue(args.force)

        mlflow_args = parse_args(
            [
                "--gold-manifest",
                "gold_manifest.json",
                "--mlflow",
                "--mlflow-tracking-uri",
                "http://mlflow.test",
                "--mlflow-experiment",
                "annotation-tests",
                "--mlflow-run-name",
                "named-run",
                "--no-mlflow-artifacts",
            ]
        )
        self.assertTrue(mlflow_args.mlflow)
        self.assertEqual(mlflow_args.mlflow_tracking_uri, "http://mlflow.test")
        self.assertEqual(mlflow_args.mlflow_experiment, "annotation-tests")
        self.assertEqual(mlflow_args.mlflow_run_name, "named-run")
        self.assertTrue(mlflow_args.no_mlflow_artifacts)

    def test_run_annotation_evaluation_scores_fixture_predictions_and_writes_artifacts(self) -> None:
        raw_prediction = {
            "entities": [
                {"type": "Drug", "name": "Fish oil", "properties": {}},
                {"type": "Biomarker", "name": "Triglycerides", "properties": {}},
            ],
            "relationships": [
                {
                    "type": "REDUCES",
                    "source": {"type": "Drug", "name": "Fish oil"},
                    "target": {"type": "Biomarker", "name": "Triglycerides"},
                    "properties": {"confidence": 0.91, "evidence": "Fish oil reduced triglycerides"},
                }
            ],
            "rejected_candidates": [],
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            gold_manifest = _write_gold_manifest(root / "gold")
            output_root = root / "eval"
            extractor = StaticFixtureExtractor({"PMC123-chunk-0001": raw_prediction}, model="fixture-model")

            with (
                mock.patch("pipelines.annotation.evaluation.get_extractor", return_value=extractor),
                mock.patch("pipelines.ingestion.neo4j_loader.load_processed_record") as load_processed_record,
            ):
                result = run_annotation_evaluation(
                    AnnotationEvaluationConfig(
                        gold_manifest_path=gold_manifest,
                        output_root=output_root,
                        eval_id="eval-fixture",
                        model_profile="noop",
                    )
                )

            self.assertEqual(result["gold_set_id"], "fixture_gold")
            self.assertEqual(result["chunk_count"], 1)
            self.assertEqual(result["success_count"], 1)
            self.assertEqual(result["metrics"]["entities"]["f1"], 1.0)
            self.assertEqual(result["metrics"]["relationships"]["f1"], 1.0)
            self.assertEqual(result["metrics"]["entity_types"]["Drug"]["f1"], 1.0)
            self.assertEqual(result["metrics"]["entity_types"]["Biomarker"]["f1"], 1.0)
            self.assertEqual(result["metrics"]["relationship_types"]["REDUCES"]["f1"], 1.0)
            self.assertTrue(result["artifact_policy"]["artifact_only"])
            self.assertEqual(result["artifact_policy"]["neo4j_load_mode"], "none")
            load_processed_record.assert_not_called()
            self.assertTrue(Path(result["eval_manifest_path"]).exists())
            self.assertTrue((output_root / "eval-fixture" / "metrics.json").exists())
            self.assertTrue((output_root / "eval-fixture" / "metrics.csv").exists())
            self.assertTrue((output_root / "eval-fixture" / "errors.csv").exists())
            self.assertTrue((output_root / "eval-fixture" / "summary.md").exists())
            self.assertTrue((output_root / "eval-fixture" / "artifact_manifest.json").exists())
            self.assertTrue((output_root / "eval-fixture" / "neo4j_load_report.json").exists())
            self.assertTrue((output_root / "eval-fixture" / "gold_snapshot" / "gold_manifest.json").exists())
            self.assertTrue((output_root / "eval-fixture" / "gold_snapshot" / "gold_entities.csv").exists())
            self.assertTrue((output_root / "eval-fixture" / "gold_snapshot" / "gold_relationships.csv").exists())
            self.assertTrue((output_root / "eval-fixture" / "gold_snapshot" / "reviewed_annotation_workbook.xlsx").exists())
            self.assertTrue((output_root / "eval-fixture" / "gold_snapshot" / "snapshot_manifest.json").exists())
            self.assertTrue((output_root / "eval-fixture" / "matches" / "entity_matches.csv").exists())
            self.assertTrue((output_root / "eval-fixture" / "predictions" / "processed" / "PMC123.json").exists())
            self.assertEqual(
                (output_root / "eval-fixture" / "errors.csv").read_text(encoding="utf-8").splitlines(),
                ["chunk_id,document_id,pmcid,stage,error"],
            )
            artifact_manifest = json.loads((output_root / "eval-fixture" / "artifact_manifest.json").read_text(encoding="utf-8"))
            artifact_paths = {item["path"] for item in artifact_manifest["artifacts"]}
            self.assertIn((output_root / "eval-fixture" / "eval_manifest.json").as_posix(), artifact_paths)
            self.assertIn((output_root / "eval-fixture" / "summary.md").as_posix(), artifact_paths)
            self.assertIn(
                (output_root / "eval-fixture" / "gold_snapshot" / "reviewed_annotation_workbook.xlsx").as_posix(),
                artifact_paths,
            )
            self.assertTrue(all(item["sha256"] for item in artifact_manifest["artifacts"]))
            neo4j_report = json.loads((output_root / "eval-fixture" / "neo4j_load_report.json").read_text(encoding="utf-8"))
            self.assertFalse(neo4j_report["enabled"])
            self.assertIn(
                "scope,label,true_positive,false_positive,false_negative,precision,recall,f1",
                (output_root / "eval-fixture" / "metrics.csv").read_text(encoding="utf-8").splitlines()[0],
            )
            self.assertIn(
                "outcome,chunk_id,entity_type,entity_name,entity_key",
                (output_root / "eval-fixture" / "matches" / "entity_matches.csv").read_text(encoding="utf-8").splitlines()[0],
            )
            self.assertIn(
                "relationship_type,source_entity_type,source_entity_name,target_entity_type,target_entity_name",
                (output_root / "eval-fixture" / "matches" / "relationship_matches.csv").read_text(encoding="utf-8").splitlines()[0],
            )

            with self.assertRaisesRegex(RuntimeError, "already exists"):
                run_annotation_evaluation(
                    AnnotationEvaluationConfig(
                        gold_manifest_path=gold_manifest,
                        output_root=output_root,
                        eval_id="eval-fixture",
                        model_profile="noop",
                    )
                )

    def test_run_annotation_evaluation_rejects_neo4j_loading_modes_without_writing_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            gold_manifest = _write_gold_manifest(root / "gold")
            output_root = root / "eval"

            with self.assertRaisesRegex(ValueError, "artifact-only"):
                run_annotation_evaluation(
                    AnnotationEvaluationConfig(
                        gold_manifest_path=gold_manifest,
                        output_root=output_root,
                        eval_id="eval-neo4j",
                        model_profile="noop",
                        neo4j_load_mode="predictions",
                    )
                )

            self.assertFalse((output_root / "eval-neo4j").exists())

    def test_run_annotation_evaluation_logs_to_mlflow_when_enabled(self) -> None:
        raw_prediction = {
            "entities": [
                {"type": "Drug", "name": "Fish oil", "properties": {}},
                {"type": "Biomarker", "name": "Triglycerides", "properties": {}},
            ],
            "relationships": [
                {
                    "type": "REDUCES",
                    "source": {"type": "Drug", "name": "Fish oil"},
                    "target": {"type": "Biomarker", "name": "Triglycerides"},
                    "properties": {"confidence": 0.91, "evidence": "Fish oil reduced triglycerides"},
                }
            ],
            "rejected_candidates": [],
        }

        calls: dict[str, list] = {
            "tracking_uri": [],
            "experiment": [],
            "start_run": [],
            "params": [],
            "metrics": [],
            "artifacts": [],
            "end_run": [],
        }

        class FakeInfo:
            run_id = "run-fixture"
            artifact_uri = "mlflow-artifacts:/run-fixture"

        class FakeRun:
            info = FakeInfo()

        fake_mlflow = types.ModuleType("mlflow")
        fake_mlflow.set_tracking_uri = lambda value: calls["tracking_uri"].append(value)
        fake_mlflow.set_experiment = lambda value: calls["experiment"].append(value)
        fake_mlflow.start_run = lambda run_name=None: calls["start_run"].append(run_name) or FakeRun()
        fake_mlflow.log_param = lambda key, value: calls["params"].append((key, value))
        fake_mlflow.log_metric = lambda key, value: calls["metrics"].append((key, value))
        fake_mlflow.log_artifacts = lambda path: calls["artifacts"].append(path)
        fake_mlflow.end_run = lambda: calls["end_run"].append(True)

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            gold_manifest = _write_gold_manifest(root / "gold")
            output_root = root / "eval"
            extractor = StaticFixtureExtractor({"PMC123-chunk-0001": raw_prediction}, model="fixture-model")

            with (
                mock.patch.dict(sys.modules, {"mlflow": fake_mlflow}),
                mock.patch("pipelines.annotation.evaluation.get_extractor", return_value=extractor),
            ):
                result = run_annotation_evaluation(
                    AnnotationEvaluationConfig(
                        gold_manifest_path=gold_manifest,
                        output_root=output_root,
                        eval_id="eval-mlflow",
                        model_profile="noop",
                        mlflow=True,
                        mlflow_tracking_uri="http://mlflow.test",
                        mlflow_experiment="annotation-tests",
                        mlflow_run_name="fixture-run",
                    )
                )

            self.assertEqual(calls["tracking_uri"], ["http://mlflow.test"])
            self.assertEqual(calls["experiment"], ["annotation-tests"])
            self.assertEqual(calls["start_run"], ["fixture-run"])
            self.assertEqual(calls["end_run"], [True])
            self.assertEqual(calls["artifacts"], [(output_root / "eval-mlflow").as_posix()])
            self.assertIn(("gold_set_id", "fixture_gold"), calls["params"])
            metric_names = {name for name, _ in calls["metrics"]}
            self.assertIn("entities_f1", metric_names)
            self.assertIn("relationships_f1", metric_names)
            self.assertIn("relationship_type_REDUCES_f1", metric_names)
            self.assertTrue(result["mlflow"]["enabled"])
            self.assertEqual(result["mlflow"]["run_id"], "run-fixture")
            self.assertEqual(result["mlflow"]["status"], "logged")

            manifest = json.loads((output_root / "eval-mlflow" / "eval_manifest.json").read_text(encoding="utf-8"))
            self.assertEqual(manifest["mlflow"]["run_id"], "run-fixture")
            self.assertEqual(manifest["mlflow"]["status"], "logged")

    def test_run_annotation_evaluation_ignores_mlflow_end_run_unicode_errors(self) -> None:
        raw_prediction = {
            "entities": [
                {"type": "Drug", "name": "Fish oil", "properties": {}},
                {"type": "Biomarker", "name": "Triglycerides", "properties": {}},
            ],
            "relationships": [
                {
                    "type": "REDUCES",
                    "source": {"type": "Drug", "name": "Fish oil"},
                    "target": {"type": "Biomarker", "name": "Triglycerides"},
                    "properties": {"confidence": 0.91, "evidence": "Fish oil reduced triglycerides"},
                }
            ],
            "rejected_candidates": [],
        }

        class FakeInfo:
            run_id = "annotation-unicode-end-run"
            artifact_uri = "mlflow-artifacts:/annotation-unicode-end-run"

        class FakeRun:
            info = FakeInfo()

        fake_mlflow = types.ModuleType("mlflow")
        fake_mlflow.set_tracking_uri = lambda value: None
        fake_mlflow.set_experiment = lambda value: None
        fake_mlflow.start_run = lambda run_name=None: FakeRun()
        fake_mlflow.log_param = lambda key, value: None
        fake_mlflow.log_metric = lambda key, value: None
        fake_mlflow.log_artifacts = lambda path: None

        def raise_unicode_error() -> None:
            raise UnicodeEncodeError("cp1252", "\U0001f3c3", 0, 1, "character maps to <undefined>")

        fake_mlflow.end_run = raise_unicode_error

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            gold_manifest = _write_gold_manifest(root / "gold")
            output_root = root / "eval"
            extractor = StaticFixtureExtractor({"PMC123-chunk-0001": raw_prediction}, model="fixture-model")

            with (
                mock.patch.dict(sys.modules, {"mlflow": fake_mlflow}),
                mock.patch("pipelines.annotation.evaluation.get_extractor", return_value=extractor),
            ):
                result = run_annotation_evaluation(
                    AnnotationEvaluationConfig(
                        gold_manifest_path=gold_manifest,
                        output_root=output_root,
                        eval_id="annotation-unicode-end-run",
                        model_profile="noop",
                        mlflow=True,
                    )
                )

        self.assertEqual(result["mlflow"]["run_id"], "annotation-unicode-end-run")
        self.assertEqual(result["mlflow"]["status"], "logged")

    def test_run_annotation_evaluation_counts_partial_tp_fp_fn_with_canonical_names(self) -> None:
        raw_prediction = {
            "entities": [
                {"type": "Drug", "name": "FISH OIL", "properties": {}},
                {"type": "Biomarker", "name": "LDL cholesterol", "properties": {}},
            ],
            "relationships": [
                {
                    "type": "REDUCES",
                    "source": {"type": "Drug", "name": "FISH OIL"},
                    "target": {"type": "Biomarker", "name": "LDL cholesterol"},
                    "properties": {"confidence": 0.91, "evidence": "Fish oil reduced triglycerides"},
                }
            ],
            "rejected_candidates": [],
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            gold_manifest = _write_gold_manifest(root / "gold")
            output_root = root / "eval"
            extractor = StaticFixtureExtractor({"PMC123-chunk-0001": raw_prediction}, model="fixture-model")

            with mock.patch("pipelines.annotation.evaluation.get_extractor", return_value=extractor):
                result = run_annotation_evaluation(
                    AnnotationEvaluationConfig(
                        gold_manifest_path=gold_manifest,
                        output_root=output_root,
                        eval_id="eval-partial",
                        model_profile="noop",
                    )
                )

            self.assertEqual(
                result["metrics"]["entities"],
                {
                    "true_positive": 1,
                    "false_positive": 1,
                    "false_negative": 1,
                    "precision": 0.5,
                    "recall": 0.5,
                    "f1": 0.5,
                },
            )
            self.assertEqual(
                result["metrics"]["relationships"],
                {
                    "true_positive": 0,
                    "false_positive": 1,
                    "false_negative": 1,
                    "precision": 0.0,
                    "recall": 0.0,
                    "f1": 0.0,
                },
            )
            self.assertEqual(result["metrics"]["entity_types"]["Drug"]["f1"], 1.0)
            self.assertEqual(result["metrics"]["entity_types"]["Biomarker"]["false_positive"], 1)
            self.assertEqual(result["metrics"]["entity_types"]["Biomarker"]["false_negative"], 1)
            self.assertEqual(result["metrics"]["relationship_types"]["REDUCES"]["false_positive"], 1)
            self.assertEqual(result["metrics"]["relationship_types"]["REDUCES"]["false_negative"], 1)

            entity_matches = (output_root / "eval-partial" / "matches" / "entity_matches.csv").read_text(encoding="utf-8")
            relationship_matches = (output_root / "eval-partial" / "matches" / "relationship_matches.csv").read_text(
                encoding="utf-8"
            )
            self.assertIn("tp,PMC123-chunk-0001,Drug,fish_oil", entity_matches)
            self.assertIn("fp,PMC123-chunk-0001,Biomarker,ldl_cholesterol", entity_matches)
            self.assertIn("fn,PMC123-chunk-0001,Biomarker,triglycerides", entity_matches)
            self.assertIn("fp,PMC123-chunk-0001,REDUCES,Drug,fish_oil,Biomarker,ldl_cholesterol", relationship_matches)
            self.assertIn("fn,PMC123-chunk-0001,REDUCES,Drug,fish_oil,Biomarker,triglycerides", relationship_matches)

    def test_run_annotation_evaluation_records_extraction_errors_as_artifacts(self) -> None:
        class FailingExtractor:
            provider = "fixture"
            model = "failing-fixture"
            last_model_call_paths: list[str] = []

            def extract(self, document: dict, chunk) -> dict:
                raise RuntimeError(f"fixture failure for {chunk.id}")

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            gold_manifest = _write_gold_manifest(root / "gold")
            output_root = root / "eval"

            with mock.patch("pipelines.annotation.evaluation.get_extractor", return_value=FailingExtractor()):
                result = run_annotation_evaluation(
                    AnnotationEvaluationConfig(
                        gold_manifest_path=gold_manifest,
                        output_root=output_root,
                        eval_id="eval-error",
                        model_profile="noop",
                    )
                )

            self.assertEqual(result["success_count"], 0)
            self.assertEqual(result["error_count"], 1)
            self.assertEqual(result["metrics"]["entities"]["false_negative"], 2)
            self.assertEqual(result["metrics"]["relationships"]["false_negative"], 1)

            errors = (output_root / "eval-error" / "errors.csv").read_text(encoding="utf-8")
            self.assertIn("PMC123-chunk-0001,paper:PMC123,PMC123,extraction,fixture failure", errors)
            processed = json.loads((output_root / "eval-error" / "predictions" / "processed" / "PMC123.json").read_text())
            self.assertEqual(processed["extractions"][0]["status"], "error")
            self.assertIn("fixture failure", processed["extractions"][0]["error"])

    def test_run_annotation_evaluation_can_skip_mlflow_artifact_upload(self) -> None:
        raw_prediction = {
            "entities": [
                {"type": "Drug", "name": "Fish oil", "properties": {}},
                {"type": "Biomarker", "name": "Triglycerides", "properties": {}},
            ],
            "relationships": [],
            "rejected_candidates": [],
        }
        calls: dict[str, list] = {"artifacts": [], "end_run": []}

        class FakeInfo:
            run_id = "run-no-artifacts"
            artifact_uri = "mlflow-artifacts:/run-no-artifacts"

        class FakeRun:
            info = FakeInfo()

        fake_mlflow = types.ModuleType("mlflow")
        fake_mlflow.set_tracking_uri = lambda value: None
        fake_mlflow.set_experiment = lambda value: None
        fake_mlflow.start_run = lambda run_name=None: FakeRun()
        fake_mlflow.log_param = lambda key, value: None
        fake_mlflow.log_metric = lambda key, value: None
        fake_mlflow.log_artifacts = lambda path: calls["artifacts"].append(path)
        fake_mlflow.end_run = lambda: calls["end_run"].append(True)

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            gold_manifest = _write_gold_manifest(root / "gold")
            output_root = root / "eval"
            extractor = StaticFixtureExtractor({"PMC123-chunk-0001": raw_prediction}, model="fixture-model")

            with (
                mock.patch.dict(sys.modules, {"mlflow": fake_mlflow}),
                mock.patch("pipelines.annotation.evaluation.get_extractor", return_value=extractor),
            ):
                result = run_annotation_evaluation(
                    AnnotationEvaluationConfig(
                        gold_manifest_path=gold_manifest,
                        output_root=output_root,
                        eval_id="eval-mlflow-no-artifacts",
                        model_profile="noop",
                        mlflow=True,
                        mlflow_log_artifacts=False,
                    )
                )

            self.assertEqual(calls["artifacts"], [])
            self.assertEqual(calls["end_run"], [True])
            self.assertTrue(result["mlflow"]["enabled"])
            self.assertFalse(result["mlflow"]["logged_artifacts"])


if __name__ == "__main__":
    unittest.main()
