from __future__ import annotations

import csv
import hashlib
import json
import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook

from pipelines.annotation.adjudicate_annotations import parse_args
from pipelines.annotation.adjudication import AdjudicationConfig, run_adjudication
from pipelines.annotation.workbook import export_annotation_workbook
from packages.llm.providers import ModelCallRecord


def _fixture_processed_record() -> dict:
    return {
        "run": {
            "id": "fixture-run",
            "created_at": "2026-01-01T00:00:00+00:00",
            "source": "pmc_bioc",
            "model_profile": "noop",
            "extractor_provider": "fixture",
            "extractor_model": "fixture-model",
            "prompt_version": "001_initial_prompt",
            "min_confidence": 0.5,
        },
        "document": {
            "id": "paper:PMC123",
            "pmcid": "PMC123",
            "pmid": "123",
            "title": "Fish oil fixture",
            "year": "2026",
            "journal": "Fixture Journal",
        },
        "chunks": [
            {
                "id": "PMC123-chunk-0001",
                "document_id": "paper:PMC123",
                "pmcid": "PMC123",
                "order": 1,
                "char_start": 100,
                "char_end": 142,
                "section": "Abstract",
                "type": "abstract",
                "text": "Fish oil reduced triglycerides in adults.",
            }
        ],
        "extractions": [
            {
                "chunk_id": "PMC123-chunk-0001",
                "status": "ok",
                "entities": [
                    {"id": "drug:fish_oil", "type": "Drug", "name": "Fish oil", "properties": {}},
                    {"id": "biomarker:triglycerides", "type": "Biomarker", "name": "Triglycerides", "properties": {}},
                ],
                "relationships": [
                    {
                        "id": "rel:abc",
                        "type": "REDUCES",
                        "source": {"id": "drug:fish_oil", "type": "Drug", "name": "Fish oil"},
                        "target": {"id": "biomarker:triglycerides", "type": "Biomarker", "name": "Triglycerides"},
                        "properties": {
                            "confidence": 0.91,
                            "evidence": "Fish oil reduced triglycerides",
                        },
                    }
                ],
                "rejected_candidates": [],
                "model_call_paths": [],
            }
        ],
    }


def _hash_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _column_index(ws, header: str) -> int:
    headers = [cell.value for cell in ws[1]]
    return headers.index(header) + 1


def _set_statuses(workbook_path: Path, *, reject_target_entity: bool = False) -> None:
    workbook = load_workbook(workbook_path)

    chunks = workbook["chunks"]
    chunk_status_col = _column_index(chunks, "annotation_status")
    for row in range(2, chunks.max_row + 1):
        chunks.cell(row=row, column=chunk_status_col).value = "reviewed"

    entities = workbook["gold_entities"]
    entity_status_col = _column_index(entities, "annotation_status")
    entity_text_col = _column_index(entities, "entity_text")
    for row in range(2, entities.max_row + 1):
        status = "accepted"
        if reject_target_entity and entities.cell(row=row, column=entity_text_col).value == "Triglycerides":
            status = "rejected"
        entities.cell(row=row, column=entity_status_col).value = status

    relationships = workbook["gold_relationships"]
    relationship_status_col = _column_index(relationships, "annotation_status")
    decision_col = _column_index(relationships, "annotation_decision")
    direction_col = _column_index(relationships, "direction_verified")
    negated_col = _column_index(relationships, "negated")
    speculative_col = _column_index(relationships, "speculative")
    for row in range(2, relationships.max_row + 1):
        relationships.cell(row=row, column=relationship_status_col).value = "accepted"
        relationships.cell(row=row, column=decision_col).value = "include"
        relationships.cell(row=row, column=direction_col).value = "yes"
        relationships.cell(row=row, column=negated_col).value = "no"
        relationships.cell(row=row, column=speculative_col).value = "no"

    workbook.save(workbook_path)


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _rows_by_header(workbook_path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        headers = [cell.value for cell in worksheet[1]]
        rows = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if any(value is not None for value in row):
                rows.append(dict(zip(headers, row)))
        return rows
    finally:
        workbook.close()


def _llm_payloads(workbook_path: Path) -> tuple[dict, dict]:
    entities = _rows_by_header(workbook_path, "gold_entities")
    relationships = _rows_by_header(workbook_path, "gold_relationships")
    entity_payload = {
        "entity_reviews": [
            {
                "entity_gold_id": row["entity_gold_id"],
                "decision": "accepted",
                "revised_entity_type": row["entity_type"],
                "revised_entity_text": row["entity_text"],
                "revised_normalized_name": row["normalized_name"],
                "evidence_text": row["evidence_text"],
                "evidence_start_char": str(row["evidence_start_char"]),
                "evidence_end_char": str(row["evidence_end_char"]),
                "reason": "Supported by fixture chunk.",
            }
            for row in entities
        ],
        "entity_additions": [],
        "review_notes": "fixture entity review",
    }
    relationship_payload = {
        "relationship_reviews": [
            {
                "relationship_gold_id": row["relationship_gold_id"],
                "decision": "accepted",
                "relationship_type": row["relationship_type"],
                "source_entity_id": row["source_entity_id"],
                "source_entity_type": row["source_entity_type"],
                "source_entity_text": row["source_entity_text"],
                "source_normalized_name": row["source_normalized_name"],
                "target_entity_id": row["target_entity_id"],
                "target_entity_type": row["target_entity_type"],
                "target_entity_text": row["target_entity_text"],
                "target_normalized_name": row["target_normalized_name"],
                "evidence_text": row["evidence_text"],
                "evidence_start_char": str(row["evidence_start_char"]),
                "evidence_end_char": str(row["evidence_end_char"]),
                "direction_verified": "yes",
                "explicit_or_implied": "explicit",
                "negated": "no",
                "speculative": "no",
                "reason": "Supported by fixture chunk.",
            }
            for row in relationships
        ],
        "relationship_additions": [],
        "review_notes": "fixture relationship review",
    }
    return entity_payload, relationship_payload


class FakeFrontierModel:
    provider = "openai"
    model = "fake-frontier-model"

    def __init__(self, entity_payload: dict, relationship_payload: dict, *, error_stage: str = "") -> None:
        self.entity_payload = entity_payload
        self.relationship_payload = relationship_payload
        self.error_stage = error_stage
        self.calls: list[str] = []

    def generate_json_record(self, prompt: str, json_schema: dict | None = None, *, prompt_version: str = "") -> ModelCallRecord:
        stage = "entity" if "ENTITY_ADJUDICATION" in prompt else "relationship"
        self.calls.append(stage)
        if self.error_stage == stage:
            return ModelCallRecord(
                provider=self.provider,
                model=self.model,
                prompt_version=prompt_version,
                request={"prompt": prompt, "json_schema": json_schema},
                json_schema=json_schema,
                response_text="",
                parsed_json={},
                raw_response={},
                started_at="2026-01-01T00:00:00+00:00",
                finished_at="2026-01-01T00:00:00+00:00",
                status="error",
                error=f"{stage} response was not valid JSON",
            )
        parsed_json = self.entity_payload if stage == "entity" else self.relationship_payload
        return ModelCallRecord(
            provider=self.provider,
            model=self.model,
            prompt_version=prompt_version,
            request={"prompt": prompt, "json_schema": json_schema},
            json_schema=json_schema,
            response_text=json.dumps(parsed_json),
            parsed_json=parsed_json,
            raw_response=parsed_json,
            started_at="2026-01-01T00:00:00+00:00",
            finished_at="2026-01-01T00:00:00+00:00",
            status="ok",
        )


class AnnotationAdjudicationTests(unittest.TestCase):
    def test_parse_args_requires_existing_workbook_path_argument_shape(self) -> None:
        args = parse_args(
            [
                "--workbook",
                "annotation_workbook.xlsx",
                "--review-id",
                "review-fixture",
                "--llm-review",
                "--model",
                "frontier-test-model",
            ]
        )
        self.assertEqual(args.workbook, Path("annotation_workbook.xlsx"))
        self.assertEqual(args.review_id, "review-fixture")
        self.assertTrue(args.llm_review)
        self.assertEqual(args.model, "frontier-test-model")

    def test_run_adjudication_exports_gold_without_mutating_original_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            workbook_path = root / "annotation_workbook.xlsx"
            export_annotation_workbook([_fixture_processed_record()], workbook_path)
            _set_statuses(workbook_path)
            before_hash = _hash_file(workbook_path)

            report = run_adjudication(
                AdjudicationConfig(
                    workbook_path=workbook_path,
                    output_root=root / "gold_v001",
                    review_id="review-valid",
                )
            )

            self.assertTrue(report["exported"])
            self.assertEqual(before_hash, _hash_file(workbook_path))
            self.assertTrue(Path(report["reviewed_workbook_path"]).exists())
            self.assertTrue(Path(report["adjudication_report_path"]).exists())
            entity_rows = _read_csv_rows(Path(report["gold_entities_path"]))
            relationship_rows = _read_csv_rows(Path(report["gold_relationships_path"]))
            self.assertEqual(len(entity_rows), 2)
            self.assertEqual(len(relationship_rows), 1)
            self.assertEqual(relationship_rows[0]["relationship_type"], "REDUCES")

    def test_run_adjudication_blocks_export_when_rows_are_unresolved(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            workbook_path = root / "annotation_workbook.xlsx"
            export_annotation_workbook([_fixture_processed_record()], workbook_path)

            report = run_adjudication(
                AdjudicationConfig(
                    workbook_path=workbook_path,
                    output_root=root / "gold_v001",
                    review_id="review-unresolved",
                )
            )

            self.assertFalse(report["exported"])
            self.assertTrue(Path(report["reviewed_workbook_path"]).exists())
            self.assertTrue(Path(report["adjudication_report_path"]).exists())
            self.assertFalse((root / "gold_v001" / "review-unresolved" / "gold_entities.csv").exists())
            self.assertGreater(report["validation"]["summary"]["unresolved_entities"], 0)
            self.assertGreater(len(report["validation"]["errors"]), 0)

    def test_run_adjudication_blocks_relationship_with_rejected_endpoint(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            workbook_path = root / "annotation_workbook.xlsx"
            export_annotation_workbook([_fixture_processed_record()], workbook_path)
            _set_statuses(workbook_path, reject_target_entity=True)

            report = run_adjudication(
                AdjudicationConfig(
                    workbook_path=workbook_path,
                    output_root=root / "gold_v001",
                    review_id="review-bad-endpoint",
                )
            )

            self.assertFalse(report["exported"])
            messages = [error["message"] for error in report["validation"]["errors"]]
            self.assertIn(
                "Accepted relationship target endpoint does not map to an accepted entity in the same chunk.",
                messages,
            )
            self.assertFalse((root / "gold_v001" / "review-bad-endpoint" / "gold_relationships.csv").exists())

    def test_llm_review_updates_copied_workbook_writes_audit_and_exports_gold(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            workbook_path = root / "annotation_workbook.xlsx"
            export_annotation_workbook([_fixture_processed_record()], workbook_path)
            entity_payload, relationship_payload = _llm_payloads(workbook_path)
            fake_model = FakeFrontierModel(entity_payload, relationship_payload)
            before_hash = _hash_file(workbook_path)

            with mock.patch("pipelines.annotation.llm_adjudication.get_language_model", return_value=fake_model):
                report = run_adjudication(
                    AdjudicationConfig(
                        workbook_path=workbook_path,
                        output_root=root / "gold_v001",
                        review_id="review-llm-valid",
                        llm_review=True,
                    )
                )

            self.assertEqual(fake_model.calls, ["entity", "relationship"])
            self.assertEqual(before_hash, _hash_file(workbook_path))
            self.assertTrue(report["exported"])
            self.assertEqual(report["llm_adjudication"]["provider"], "openai")
            self.assertEqual(report["llm_adjudication"]["model"], fake_model.model)
            self.assertEqual(report["llm_adjudication"]["reviewed_chunk_count"], 1)
            self.assertEqual(report["llm_adjudication"]["failed_chunk_count"], 0)
            self.assertEqual(len(report["llm_adjudication"]["model_call_paths"]), 2)
            for path in report["llm_adjudication"]["model_call_paths"]:
                self.assertTrue(Path(path).exists())

            copied_entities = _rows_by_header(Path(report["reviewed_workbook_path"]), "gold_entities")
            copied_relationships = _rows_by_header(Path(report["reviewed_workbook_path"]), "gold_relationships")
            copied_chunks = _rows_by_header(Path(report["reviewed_workbook_path"]), "chunks")
            self.assertEqual({row["annotation_status"] for row in copied_entities}, {"accepted"})
            self.assertEqual({row["annotation_status"] for row in copied_relationships}, {"accepted"})
            self.assertEqual(copied_relationships[0]["annotation_decision"], "include")
            self.assertEqual(copied_relationships[0]["direction_verified"], "yes")
            self.assertEqual(copied_chunks[0]["annotation_status"], "reviewed")
            self.assertEqual(len(_read_csv_rows(Path(report["gold_entities_path"]))), 2)
            self.assertEqual(len(_read_csv_rows(Path(report["gold_relationships_path"]))), 1)

    def test_llm_review_rejects_non_frontier_profile(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            workbook_path = root / "annotation_workbook.xlsx"
            export_annotation_workbook([_fixture_processed_record()], workbook_path)

            with self.assertRaisesRegex(ValueError, "frontier OpenAI"):
                run_adjudication(
                    AdjudicationConfig(
                        workbook_path=workbook_path,
                        output_root=root / "gold_v001",
                        review_id="review-invalid-profile",
                        llm_review=True,
                        model_profile="noop",
                    )
                )

    def test_llm_review_model_error_blocks_export_and_records_chunk_error(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            workbook_path = root / "annotation_workbook.xlsx"
            export_annotation_workbook([_fixture_processed_record()], workbook_path)
            entity_payload, relationship_payload = _llm_payloads(workbook_path)
            fake_model = FakeFrontierModel(entity_payload, relationship_payload, error_stage="entity")

            with mock.patch("pipelines.annotation.llm_adjudication.get_language_model", return_value=fake_model):
                report = run_adjudication(
                    AdjudicationConfig(
                        workbook_path=workbook_path,
                        output_root=root / "gold_v001",
                        review_id="review-llm-error",
                        llm_review=True,
                    )
                )

            self.assertFalse(report["exported"])
            self.assertEqual(fake_model.calls, ["entity"])
            self.assertEqual(report["llm_adjudication"]["reviewed_chunk_count"], 0)
            self.assertEqual(report["llm_adjudication"]["failed_chunk_count"], 1)
            self.assertIn("not valid JSON", report["llm_adjudication"]["chunks"][0]["error"])
            self.assertEqual(len(report["llm_adjudication"]["model_call_paths"]), 1)
            self.assertTrue(Path(report["llm_adjudication"]["model_call_paths"][0]).exists())
            self.assertFalse((root / "gold_v001" / "review-llm-error" / "gold_entities.csv").exists())


if __name__ == "__main__":
    unittest.main()
